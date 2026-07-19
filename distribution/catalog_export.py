import csv
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from zipfile import ZIP_DEFLATED, ZipFile

from django.db.models import Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    AcquisitionAgreement,
    CinemaBooking,
    Cost,
    ExploitationField,
    RightsWindow,
    RoyaltyStatement,
    SalesAgreement,
    SalesReport,
    Title,
    TitleMaterial,
    WaterfallPlan,
    WaterfallRun,
    WaterfallRunCostAllocation,
    WaterfallRunLine,
    WaterfallRunStatus,
    WaterfallStep,
)


@dataclass(frozen=True)
class ExportSheet:
    name: str
    csv_name: str
    headers: list[str]
    rows: list[list]


def _yes(value):
    return "tak" if value else "nie"


def _join(values):
    return ", ".join(str(value) for value in values if value)


def _file_name(field):
    return field.name if field else ""


def _period_filter(queryset, start_field, end_field, date_from, date_to):
    if date_from and date_to:
        return queryset.filter(**{f"{start_field}__lte": date_to, f"{end_field}__gte": date_from})
    return queryset


def _cost_recovered(cost):
    return cost.recovered_total or Decimal("0.00")


def _cost_outstanding(cost):
    if not cost.recoupable:
        return Decimal("0.00")
    return max(cost.net_amount - _cost_recovered(cost), Decimal("0.00"))


def build_catalog_export(title_ids, *, date_from=None, date_to=None):
    titles = list(
        Title.objects.filter(pk__in=title_ids)
        .select_related("producer")
        .order_by("title_pl", "pk")
    )
    selected_ids = [title.pk for title in titles]
    field_labels = dict(ExploitationField.choices)

    acquisitions = list(
        AcquisitionAgreement.objects.filter(title_id__in=selected_ids)
        .select_related("title", "licensor")
        .prefetch_related("territories")
        .order_by("title__title_pl", "contract_number", "pk")
    )
    sales_agreements = list(
        SalesAgreement.objects.filter(title_id__in=selected_ids)
        .select_related("title", "licensee")
        .order_by("title__title_pl", "contract_number", "pk")
    )
    rights = list(
        RightsWindow.objects.filter(title_id__in=selected_ids)
        .select_related("title", "counterparty", "acquisition_agreement", "sales_agreement")
        .prefetch_related("territories", "language_versions")
        .order_by("title__title_pl", "exploitation_field", "date_from")
    )

    sales_reports_qs = SalesReport.objects.filter(title_id__in=selected_ids).select_related(
        "title", "counterparty", "territory", "sales_agreement"
    )
    sales_reports_qs = _period_filter(sales_reports_qs, "period_start", "period_end", date_from, date_to)
    sales_reports = list(sales_reports_qs.order_by("title__title_pl", "period_start", "pk"))

    bookings_qs = CinemaBooking.objects.filter(title_id__in=selected_ids).select_related("title", "cinema")
    bookings_qs = _period_filter(bookings_qs, "date_from", "date_to", date_from, date_to)
    bookings = list(bookings_qs.order_by("title__title_pl", "date_from", "pk"))

    costs_qs = (
        Cost.objects.filter(title_id__in=selected_ids)
        .select_related("title", "supplier")
        .annotate(
            recovered_total=Sum(
                "waterfall_allocations__allocated_amount",
                filter=Q(waterfall_allocations__run_line__run__status=WaterfallRunStatus.FINALIZED),
            )
        )
    )
    if date_from and date_to:
        costs_qs = costs_qs.filter(cost_date__range=(date_from, date_to))
    costs = list(costs_qs.order_by("title__title_pl", "cost_date", "pk"))

    materials = list(
        TitleMaterial.objects.filter(title_id__in=selected_ids)
        .select_related("title", "language_version", "supplier")
        .order_by("title__title_pl", "asset_type", "pk")
    )
    plans = list(WaterfallPlan.objects.filter(title_id__in=selected_ids).select_related("title").order_by("title__title_pl", "version"))
    plan_ids = [plan.pk for plan in plans]
    steps = list(
        WaterfallStep.objects.filter(plan_id__in=plan_ids)
        .select_related("plan", "plan__title", "beneficiary")
        .order_by("plan__title__title_pl", "plan__version", "phase", "sort_order", "pk")
    )
    runs_qs = WaterfallRun.objects.filter(plan_id__in=plan_ids).select_related("plan", "plan__title", "finalized_by")
    runs_qs = _period_filter(runs_qs, "period_start", "period_end", date_from, date_to)
    runs = list(runs_qs.order_by("plan__title__title_pl", "period_start", "pk"))
    run_ids = [run.pk for run in runs]
    lines = list(
        WaterfallRunLine.objects.filter(run_id__in=run_ids)
        .select_related("run", "run__plan", "run__plan__title", "step", "beneficiary")
        .order_by("run__plan__title__title_pl", "run__period_start", "sequence")
    )
    allocations = list(
        WaterfallRunCostAllocation.objects.filter(run_line__run_id__in=run_ids)
        .select_related("run_line", "run_line__run", "run_line__run__plan", "run_line__run__plan__title", "cost")
        .order_by("run_line__run__plan__title__title_pl", "run_line__run__period_start", "run_line__sequence", "cost_id")
    )

    statements_qs = RoyaltyStatement.objects.filter(title_id__in=selected_ids).select_related(
        "title", "recipient", "waterfall_plan", "waterfall_run"
    )
    statements_qs = _period_filter(statements_qs, "period_start", "period_end", date_from, date_to)
    statements = list(statements_qs.order_by("title__title_pl", "period_start", "recipient__name"))

    acquisition_count = defaultdict(int)
    rights_count = defaultdict(int)
    material_count = defaultdict(int)
    for agreement in acquisitions:
        acquisition_count[agreement.title_id] += 1
    for right in rights:
        rights_count[right.title_id] += 1
    for material in materials:
        material_count[material.title_id] += 1

    summary_keys = {(title.pk, title.acquisition_currency) for title in titles}
    for item in sales_reports:
        summary_keys.add((item.title_id, item.currency))
    for item in costs:
        summary_keys.add((item.title_id, item.currency))
    for item in statements:
        summary_keys.add((item.title_id, item.currency))
    for item in runs:
        summary_keys.add((item.plan.title_id, item.plan.currency))

    titles_by_id = {title.pk: title for title in titles}
    summary_rows = []
    for title_id, currency in sorted(summary_keys, key=lambda key: (titles_by_id[key[0]].title_pl.lower(), key[1])):
        title_sales = [item for item in sales_reports if item.title_id == title_id and item.currency == currency]
        title_costs = [item for item in costs if item.title_id == title_id and item.currency == currency]
        title_statements = [item for item in statements if item.title_id == title_id and item.currency == currency]
        title_allocations = [
            item for item in allocations
            if item.run_line.run.plan.title_id == title_id
            and item.run_line.run.plan.currency == currency
            and item.run_line.run.status == WaterfallRunStatus.FINALIZED
        ]
        summary_rows.append([
            title_id,
            titles_by_id[title_id].title_pl,
            currency,
            sum((item.gross_revenue for item in title_sales), Decimal("0.00")),
            sum((item.net_revenue for item in title_sales), Decimal("0.00")),
            sum((item.net_amount for item in title_costs), Decimal("0.00")),
            sum((item.net_amount for item in title_costs if item.recoupable), Decimal("0.00")),
            sum((item.allocated_amount for item in title_allocations), Decimal("0.00")),
            sum((item.amount_due for item in title_statements), Decimal("0.00")),
            len(title_sales),
            len(title_costs),
            len(title_statements),
            acquisition_count[title_id],
            rights_count[title_id],
            material_count[title_id],
        ])

    sheets = [
        ExportSheet(
            "Zakres eksportu",
            "00_zakres_eksportu.csv",
            ["Parametr", "Wartość"],
            [
                ["Wygenerowano", timezone.localtime().replace(tzinfo=None)],
                ["Liczba tytułów", len(titles)],
                ["Zakres finansowy", "wybrany okres" if date_from else "cała historia"],
                ["Data od", date_from or ""],
                ["Data do", date_to or ""],
            ],
        ),
        ExportSheet(
            "Podsumowanie",
            "01_podsumowanie.csv",
            ["ID tytułu", "Tytuł", "Waluta", "Przychód brutto", "Przychód netto", "Koszty netto", "Koszty recoupable", "Odzyskano w rozliczeniach", "Statementy do wypłaty", "Liczba raportów", "Liczba kosztów", "Liczba statementów", "Umowy nabycia", "Okna praw", "Materiały"],
            summary_rows,
        ),
        ExportSheet(
            "Tytuły",
            "02_tytuly.csv",
            ["ID", "Tytuł PL", "Tytuł oryginalny", "Rok", "Kraje", "Czas min", "Status", "Premiera PL", "Producent", "Waluta nabycia", "MG", "IMDb", "EAN", "Nośnik", "Kategoria marketplace", "ID kategorii", "Stan", "Wydanie", "Opakowanie", "Liczba nośników", "Region", "Gatunek", "Reżyser", "Obsada", "Scenariusz", "Muzyka", "Audio", "Napisy", "Dubbing", "Lektor", "Wiek", "Kolor", "Format obrazu", "Opis marketplace", "Tagi", "Uwagi"],
            [[title.pk, title.title_pl, title.original_title, title.production_year or "", title.countries, title.runtime_minutes or "", title.get_status_display(), title.polish_premiere_date or "", title.producer.name if title.producer else "", title.acquisition_currency, title.mg_advance, title.imdb_url, title.ean, title.get_media_type_display(), title.marketplace_category_name, title.marketplace_category_id, title.get_marketplace_condition_display(), title.release_edition, title.package_type, title.discs_count or "", title.region_code, title.genre, title.director, title.cast, title.screenwriter, title.music_by, title.audio_languages, title.subtitle_languages, title.dubbing_languages, title.lector_languages, title.get_age_rating_display(), title.color_mode, title.aspect_ratio, title.marketplace_description, title.marketplace_tags, title.notes] for title in titles],
        ),
        ExportSheet(
            "Umowy nabycia",
            "03_umowy_nabycia.csv",
            ["ID", "Tytuł", "Numer umowy", "Licencjodawca", "Podpisano", "Prawa od", "Prawa do", "Terytoria", "Waluta", "MG", "Revenue share %", "P&A recoupable", "Status", "Plik", "Uwagi"],
            [[item.pk, item.title.title_pl, item.contract_number, item.licensor.name, item.signed_date or "", item.rights_start or "", item.rights_end or "", _join(item.territories.all()), item.currency, item.mg_advance, item.revenue_share_percent, _yes(item.pa_recoupable), item.get_status_display(), _file_name(item.agreement_file), item.notes] for item in acquisitions],
        ),
        ExportSheet(
            "Umowy sprzedaży",
            "04_umowy_sprzedazy.csv",
            ["ID", "Tytuł", "Numer licencji", "Licencjobiorca", "Podpisano", "Status", "Waluta", "Fixed fee", "Revenue share %", "Raportowanie", "Termin płatności", "Faktura wystawiona", "Faktura zapłacona", "Plik", "Uwagi"],
            [[item.pk, item.title.title_pl, item.contract_number, item.licensee.name, item.signed_date or "", item.get_status_display(), item.currency, item.fixed_fee, item.revenue_share_percent, item.get_reporting_cycle_display(), item.payment_due_date or "", _yes(item.invoice_issued), _yes(item.invoice_paid), _file_name(item.agreement_file), item.notes] for item in sales_agreements],
        ),
        ExportSheet(
            "Prawa",
            "05_prawa.csv",
            ["ID", "Tytuł", "Źródło", "Umowa nabycia", "Umowa sprzedaży", "Kontrahent", "Pole eksploatacji", "Terytoria", "Wersje językowe", "Od", "Do", "Wyłączność", "Holdback", "Status", "Konflikty", "Uwagi"],
            [[item.pk, item.title.title_pl, item.get_source_display(), item.acquisition_agreement.contract_number if item.acquisition_agreement else "", item.sales_agreement.contract_number if item.sales_agreement else "", item.counterparty.name if item.counterparty else "", item.get_exploitation_field_display(), _join(item.territories.all()), _join(item.language_versions.all()), item.date_from, item.date_to, _yes(item.exclusive), _yes(item.holdback), item.get_status_display(), item.conflict_notes, item.notes] for item in rights],
        ),
        ExportSheet(
            "Raporty sprzedaży",
            "06_raporty_sprzedazy.csv",
            ["ID", "Tytuł", "Kontrahent", "Umowa", "Pole eksploatacji", "Terytorium", "Okres od", "Okres do", "Waluta", "Brutto", "Potrącenia", "VAT withholding", "Netto", "Status", "Referencja", "Plik", "Uwagi"],
            [[item.pk, item.title.title_pl, item.counterparty.name, item.sales_agreement.contract_number if item.sales_agreement else "", item.get_exploitation_field_display(), item.territory.name if item.territory else "", item.period_start, item.period_end, item.currency, item.gross_revenue, item.deductions, item.vat_withholding, item.net_revenue, item.get_status_display(), item.source_reference, _file_name(item.source_file), item.notes] for item in sales_reports],
        ),
        ExportSheet(
            "Seanse kinowe",
            "07_seanse_kinowe.csv",
            ["ID", "Tytuł", "Kino", "Miasto", "Od", "Do", "Tydzień eksploatacji", "Seanse", "Widzowie", "Box office brutto", "Udział dystrybutora %", "Faktura", "Referencja importu", "Plik", "Uwagi"],
            [[item.pk, item.title.title_pl, item.cinema.name, item.city, item.date_from, item.date_to, item.exploitation_week or "", item.screenings, item.admissions, item.box_office_gross, item.distributor_share_percent, _yes(item.invoice_issued), item.source_reference, _file_name(item.source_file), item.notes] for item in bookings],
        ),
        ExportSheet(
            "Koszty P&A",
            "08_koszty_pa.csv",
            ["ID", "Tytuł", "Dostawca", "Kategoria", "Data", "Waluta", "Netto", "VAT", "Brutto", "Recoupable", "Tryb zakresu", "Zakres", "Podział procentowy", "Odzyskano łącznie", "Pozostało do odzyskania", "Zapłacono", "Faktura", "Uwagi"],
            [[item.pk, item.title.title_pl, item.supplier.name if item.supplier else "", item.get_category_display(), item.cost_date, item.currency, item.net_amount, item.vat_amount, item.gross_amount, _yes(item.recoupable), item.get_scope_mode_display(), item.scope_label, _join(f"{field_labels.get(field, field)} {percent}%" for field, percent in item.allocation_percentages.items()), _cost_recovered(item), _cost_outstanding(item), _yes(item.paid), _file_name(item.invoice_file), item.notes] for item in costs],
        ),
        ExportSheet(
            "Materiały",
            "09_materialy.csv",
            ["ID", "Tytuł", "Typ", "Pole eksploatacji", "Wersja językowa", "Dostawca", "Status", "Wymagane", "Termin", "Dostarczono", "Plik", "Referencja", "Uwagi"],
            [[item.pk, item.title.title_pl, item.get_asset_type_display(), item.get_exploitation_field_display() if item.exploitation_field else "ogólne", item.language_version.name if item.language_version else "", item.supplier.name if item.supplier else "", item.get_status_display(), _yes(item.required_for_release), item.due_date or "", item.delivered_at or "", _file_name(item.file), item.external_reference, item.notes] for item in materials],
        ),
        ExportSheet(
            "Plany waterfall",
            "10_plany_waterfall.csv",
            ["ID", "Tytuł", "Nazwa", "Wersja", "Status", "Waluta", "Obowiązuje od", "Obowiązuje do", "Wszystkie pola", "Pola eksploatacji", "Uwagi"],
            [[item.pk, item.title.title_pl, item.name, item.version, item.get_status_display(), item.currency, item.effective_from or "", item.effective_to or "", _yes(item.applies_to_all_exploitation_fields), _join(field_labels.get(field, field) for field in item.exploitation_fields), item.notes] for item in plans],
        ),
        ExportSheet(
            "Kroki waterfall",
            "11_kroki_waterfall.csv",
            ["ID", "Tytuł", "Plan", "Wersja", "Faza", "Kolejność", "Nazwa", "Typ", "Tryb alokacji", "Beneficjent", "Procent", "Kwota stała", "Cel recoupment", "Uplift %", "Odzyskano wcześniej", "Cap", "Dolicz MG", "Dolicz koszty", "Kategorie kosztów", "Pola eksploatacji", "Aktywny", "Podstawa umowna"],
            [[item.pk, item.plan.title.title_pl, item.plan.name, item.plan.version, item.phase, item.sort_order, item.name, item.get_step_type_display(), item.get_allocation_mode_display(), item.beneficiary.name if item.beneficiary else "", item.percentage, item.fixed_amount, item.target_amount, item.premium_percent, item.opening_recouped, item.cap_amount, _yes(item.include_title_mg), _yes(item.include_recoupable_costs), _join(item.cost_categories), _join(field_labels.get(field, field) for field in item.exploitation_fields), _yes(item.active), item.notes] for item in steps],
        ),
        ExportSheet(
            "Rozliczenia waterfall",
            "12_rozliczenia_waterfall.csv",
            ["ID", "Tytuł", "Plan", "Wersja", "Okres od", "Okres do", "Status", "Waluta", "Przychód brutto", "Przychód netto", "Rozdzielono", "Pozostało", "Przeliczono", "Zatwierdzono", "Zatwierdził", "Uwagi"],
            [[item.pk, item.plan.title.title_pl, item.plan.name, item.plan.version, item.period_start, item.period_end, item.get_status_display(), item.plan.currency, item.gross_revenue, item.net_revenue, item.allocated_amount, item.closing_available, item.calculated_at.replace(tzinfo=None) if item.calculated_at else "", item.finalized_at.replace(tzinfo=None) if item.finalized_at else "", item.finalized_by.get_username() if item.finalized_by else "", item.notes] for item in runs],
        ),
        ExportSheet(
            "Pozycje waterfall",
            "13_pozycje_waterfall.csv",
            ["ID", "ID rozliczenia", "Tytuł", "Okres od", "Okres do", "Sekwencja", "Faza", "Krok", "Beneficjent", "Dostępne przed", "Podstawa", "Alokacja", "Dostępne po", "Recoupment przed", "Recoupment po", "Szczegóły"],
            [[item.pk, item.run_id, item.run.plan.title.title_pl, item.run.period_start, item.run.period_end, item.sequence, item.phase, item.step.name, item.beneficiary.name if item.beneficiary else "", item.opening_available, item.calculation_base, item.allocated_amount, item.closing_available, item.opening_recoupment, item.closing_recoupment, json.dumps(item.calculation_details, ensure_ascii=False, sort_keys=True)] for item in lines],
        ),
        ExportSheet(
            "Alokacje kosztów",
            "14_alokacje_kosztow.csv",
            ["ID", "ID rozliczenia", "Tytuł", "Okres od", "Okres do", "Status rozliczenia", "Krok", "ID kosztu", "Data kosztu", "Kategoria", "Część pola", "Odzyskana kwota", "Waluta"],
            [[item.pk, item.run_line.run_id, item.run_line.run.plan.title.title_pl, item.run_line.run.period_start, item.run_line.run.period_end, item.run_line.run.get_status_display(), item.run_line.step.name, item.cost_id, item.cost.cost_date, item.cost.get_category_display(), field_labels.get(item.exploitation_field, item.exploitation_field) if item.exploitation_field else "wspólna pula", item.allocated_amount, item.cost.currency] for item in allocations],
        ),
        ExportSheet(
            "Statementy",
            "15_statementy.csv",
            ["ID", "Tytuł", "Odbiorca", "Okres od", "Okres do", "Waluta", "Plan", "Rozliczenie waterfall ID", "Przychód brutto", "Przychód netto", "Koszty odzyskane", "Fee dystrybutora", "Net receipts", "Do wypłaty", "Status", "Wysłano", "Zapłacono", "PDF", "Uwagi"],
            [[item.pk, item.title.title_pl, item.recipient.name, item.period_start, item.period_end, item.currency, item.waterfall_plan.name if item.waterfall_plan else "", item.waterfall_run_id or "", item.gross_revenue, item.net_revenue, item.recoupable_costs, item.distributor_fee_amount, item.net_receipts, item.amount_due, item.get_status_display(), item.sent_at or "", item.paid_at or "", _file_name(item.statement_file), item.notes] for item in statements],
        ),
    ]
    return sheets


def _write_xlsx_sheet(sheet, export_sheet):
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(export_sheet.headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0058F8")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30

    for row in export_sheet.rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, Decimal):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(cell.value, str) and len(cell.value) > 40)

    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        maximum = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(maximum + 2, 12), 45)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def export_catalog_xlsx(sheets):
    workbook = Workbook()
    default = workbook.active
    for index, export_sheet in enumerate(sheets):
        sheet = default if index == 0 else workbook.create_sheet()
        sheet.title = export_sheet.name[:31]
        _write_xlsx_sheet(sheet, export_sheet)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _csv_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return str(value) if value is not None else ""


def export_catalog_csv_zip(sheets):
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for export_sheet in sheets:
            text = StringIO(newline="")
            writer = csv.writer(text, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
            writer.writerow(export_sheet.headers)
            writer.writerows([[_csv_value(value) for value in row] for row in export_sheet.rows])
            archive.writestr(export_sheet.csv_name, text.getvalue().encode("utf-8-sig"))
    return output.getvalue()
