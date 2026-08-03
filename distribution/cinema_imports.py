import hashlib
import re
import unicodedata
from difflib import SequenceMatcher
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pdfplumber
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    CinemaBookingWeek,
    CinemaReportImport,
    CinemaReportImportRow,
    Counterparty,
    CounterpartyType,
    ImportStatus,
    Title,
)


DATE_RE = re.compile(r"\b(\d{4}-\d{2}-\d{2}|\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\b")
MONEY_RE = re.compile(r"(?<![\d.])(\d{1,3}(?:[ \u00a0]?\d{3})*[,.]\d{2})(?!\d)")


def ensure_file_hash(report_import):
    digest = hashlib.sha256()
    report_import.source_file.open("rb")
    try:
        for chunk in report_import.source_file.chunks():
            digest.update(chunk)
    finally:
        report_import.source_file.close()
    file_hash = digest.hexdigest()
    if report_import.file_hash != file_hash:
        report_import.file_hash = file_hash
        report_import.save(update_fields=["file_hash", "updated_at"])
    return file_hash


def source_fingerprint(report_import, identity):
    file_hash = report_import.file_hash or ensure_file_hash(report_import)
    normalized_identity = normalize_text(identity)
    return hashlib.sha256(f"{file_hash}|{normalized_identity}".encode("utf-8")).hexdigest()


def upsert_import_row(report_import, identity, defaults):
    fingerprint = source_fingerprint(report_import, identity)
    existing = report_import.rows.filter(source_fingerprint=fingerprint).first()
    if existing and existing.status == ImportStatus.IMPORTED:
        return existing
    if existing:
        for field_name, value in defaults.items():
            setattr(existing, field_name, value)
        existing.save(update_fields=[*defaults.keys(), "updated_at"])
        return existing
    return CinemaReportImportRow.objects.create(
        report_import=report_import,
        source_fingerprint=fingerprint,
        **defaults,
    )


def parse_date(value):
    value = str(value or "").strip()
    for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y", "%d-%m-%y"]:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def normalize_date_pair(date_from, date_to):
    if date_from and date_to and date_from > date_to:
        return date_to, date_from
    return date_from, date_to


def parse_date_range(value):
    value = str(value or "").strip()
    if not value:
        return None, None
    direct = parse_date(value)
    if direct:
        return direct, direct
    match = re.search(r"(\d{1,2})\s*[-–]\s*(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", value)
    if match:
        start_day, end_day, month, year = match.groups()
        year = f"20{year}" if len(year) == 2 else year
        start = parse_date(f"{start_day}.{month}.{year}")
        end = parse_date(f"{end_day}.{month}.{year}")
        return normalize_date_pair(start, end)
    dates = [parse_date(item) for item in DATE_RE.findall(value)]
    dates = [date for date in dates if date]
    if dates:
        return normalize_date_pair(dates[0], dates[-1])
    return None, None


def parse_decimal(value):
    cleaned = str(value or "0").replace(" ", "").replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0.00")


def parse_int(value):
    if isinstance(value, (int, float)):
        return int(value)
    cleaned = re.sub(r"\D", "", str(value or ""))
    return int(cleaned) if cleaned else 0


def normalize_text(value):
    normalized = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", normalized.lower()).strip()


def extract_pdf_text(path):
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    return "\n".join(pages)


def find_known_title(text):
    lowered = normalize_text(text)
    lowered_words = lowered.split()
    best_title = None
    best_ratio = 0
    for title in Title.objects.order_by("-title_pl"):
        candidates = [normalize_text(title.title_pl), normalize_text(title.original_title)]
        if any(candidate and candidate in lowered for candidate in candidates):
            return title
        for candidate in candidates:
            if not candidate:
                continue
            candidate_words = candidate.split()
            windows = [" ".join(lowered_words[i:i + len(candidate_words)]) for i in range(max(len(lowered_words) - len(candidate_words) + 1, 1))]
            ratio = max([SequenceMatcher(None, candidate, window).ratio() for window in windows] or [0])
            if ratio > best_ratio:
                best_ratio = ratio
                best_title = title
    if best_ratio >= 0.78:
        return best_title
    return None


def find_known_title_from_import(report_import, extra_text=""):
    candidates = [
        report_import.original_filename,
        Path(report_import.source_file.name).stem,
        str(report_import.source_file.name).replace("_", " "),
        extra_text,
    ]
    for candidate in candidates:
        title = find_known_title(candidate)
        if title:
            return title
    return None


def find_known_cinema(text):
    lowered = normalize_text(text)
    qs = Counterparty.objects.filter(counterparty_type__in=[CounterpartyType.CINEMA, CounterpartyType.CINEMA_CHAIN])
    for counterparty in qs.order_by("-name"):
        if normalize_text(counterparty.name) in lowered:
            return counterparty
    for counterparty in Counterparty.objects.order_by("-name"):
        if normalize_text(counterparty.name) in lowered:
            return counterparty
    return None


def get_or_create_cinema(name):
    name = str(name or "").strip()
    if not name:
        return None
    found = find_known_cinema(name)
    if found:
        return found
    cinema, _ = Counterparty.objects.get_or_create(
        name=name,
        defaults={"counterparty_type": CounterpartyType.CINEMA},
    )
    return cinema


def infer_row_from_line(line, fallback_title=None, fallback_cinema=None):
    dates = [parse_date(match.group(1)) for match in DATE_RE.finditer(line)]
    dates = [date for date in dates if date]
    line_without_dates = DATE_RE.sub(" ", line)
    money_matches = [match.group(1) for match in MONEY_RE.finditer(line_without_dates)]
    line_without_money = MONEY_RE.sub(" ", line_without_dates)
    numbers = [parse_int(match.group(1)) for match in re.finditer(r"(?<!\d)(\d{1,6})(?!\d)", line_without_money)]
    title = find_known_title(line) or fallback_title
    cinema = find_known_cinema(line) or fallback_cinema
    confidence = Decimal("20.00")
    if title:
        confidence += Decimal("25.00")
    if cinema:
        confidence += Decimal("20.00")
    if dates:
        confidence += Decimal("20.00")
    if money_matches:
        confidence += Decimal("15.00")

    screenings = numbers[0] if len(numbers) >= 2 else 0
    admissions = numbers[1] if len(numbers) >= 2 else (numbers[0] if numbers else 0)
    normalized_line = normalize_text(line)
    rental_labels = ("film rental", "naleznosc dystrybutora", "udzial dystrybutora", "do zaplaty")
    has_rental_label = any(label in normalized_line for label in rental_labels)
    reported_rental = parse_decimal(money_matches[-1]) if has_rental_label and len(money_matches) >= 2 else None
    box_office = parse_decimal(money_matches[-2]) if reported_rental is not None else (
        parse_decimal(money_matches[-1]) if money_matches else Decimal("0.00")
    )
    date_from = dates[0] if dates else None
    date_to = dates[-1] if dates else date_from
    date_from, date_to = normalize_date_pair(date_from, date_to)
    return {
        "title": title,
        "cinema": cinema,
        "date_from": date_from,
        "date_to": date_to,
        "screenings": screenings,
        "admissions": admissions,
        "box_office_gross": box_office,
        "reported_rental_amount": reported_rental,
        "confidence": min(confidence, Decimal("100.00")),
        "source_line": line,
        "raw_payload": {"dates": [str(date) for date in dates], "numbers": numbers, "money": money_matches},
    }


def parse_cinema_report_pdf(report_import):
    ensure_file_hash(report_import)
    path = report_import.source_file.path
    text = extract_pdf_text(path)
    fallback_title = find_known_title(text)
    fallback_cinema = find_known_cinema(text)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    candidate_lines = [line for line in lines if DATE_RE.search(line) or MONEY_RE.search(line)]

    report_import.rows.exclude(status=ImportStatus.IMPORTED).delete()
    created = 0
    for line_number, line in enumerate(candidate_lines, start=1):
        inferred = infer_row_from_line(line, fallback_title=fallback_title, fallback_cinema=fallback_cinema)
        if not any([inferred["date_from"], inferred["box_office_gross"], inferred["admissions"], inferred["title"], inferred["cinema"]]):
            continue
        upsert_import_row(
            report_import,
            f"pdf-line-{line_number}|{line}",
            {"status": ImportStatus.NEEDS_REVIEW, **inferred},
        )
        created += 1

    if created == 0:
        upsert_import_row(
            report_import,
            "pdf-fallback",
            {
                "status": ImportStatus.NEEDS_REVIEW,
                "title": fallback_title,
                "cinema": fallback_cinema,
                "confidence": Decimal("10.00"),
                "source_line": text[:4000],
                "raw_payload": {"full_text": text[:10000]},
                "notes": "Nie udało się rozbić PDF na wiersze. Uzupełnij dane ręcznie przed akceptacją.",
            },
        )
        created = 1

    report_import.original_filename = report_import.original_filename or Path(report_import.source_file.name).name
    report_import.status = ImportStatus.NEEDS_REVIEW
    report_import.parsed_at = timezone.now()
    duplicate = CinemaReportImport.objects.filter(
        file_hash=report_import.file_hash,
        status=ImportStatus.IMPORTED,
    ).exclude(pk=report_import.pk).first()
    duplicate_note = f" Ten sam plik był już zaimportowany jako #{duplicate.pk}." if duplicate else ""
    report_import.parser_notes = f"Rozpoznano {created} wierszy z PDF. Sprawdź dane przed akceptacją.{duplicate_note}"
    report_import.save(update_fields=["original_filename", "status", "parsed_at", "parser_notes", "updated_at"])
    return created


def _xlsx_cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _find_metric_columns(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True))
    metric_columns = []
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if normalize_text(value) in {"widzow", "widzowie", "liczba widzow"}:
                date_value = None
                for lookup_row in range(row_index - 1, 0, -1):
                    candidate = ws.cell(lookup_row, col_index).value
                    if candidate:
                        date_value = candidate
                        break
                date_from, date_to = parse_date_range(date_value)
                metric_columns.append(
                    {
                        "row": row_index,
                        "admissions_col": col_index,
                        "gross_col": col_index + 1,
                        "net_col": col_index + 2,
                        "rental_col": next(
                            (
                                candidate_col
                                for candidate_col in range(col_index + 1, min(col_index + 8, ws.max_column) + 1)
                                if normalize_text(ws.cell(row_index, candidate_col).value)
                                in {
                                    "film rental",
                                    "naleznosc dystrybutora",
                                    "udzial dystrybutora",
                                    "kwota dystrybutora",
                                    "do zaplaty dystrybutorowi",
                                }
                            ),
                            None,
                        ),
                        "date_from": date_from,
                        "date_to": date_to,
                        "date_label": _xlsx_cell_text(date_value),
                    }
                )
    return metric_columns


def _find_identity_columns(ws, metric_header_row):
    sample_rows = list(ws.iter_rows(min_row=metric_header_row + 1, max_row=min(ws.max_row, metric_header_row + 25), values_only=True))
    best_pair = (3, 4)
    best_score = 0
    for city_col in range(1, min(ws.max_column, 8)):
        cinema_col = city_col + 1
        score = 0
        for row in sample_rows:
            city = row[city_col - 1] if len(row) >= city_col else None
            cinema = row[cinema_col - 1] if len(row) >= cinema_col else None
            if not isinstance(city, str) or not city.strip():
                continue
            if not isinstance(cinema, str) or not cinema.strip():
                continue
            city_len = len(city.strip())
            cinema_len = len(cinema.strip())
            if city_len <= 35 and cinema_len <= 60:
                score += 3
            elif city_len <= 50 and cinema_len <= 80:
                score += 1
        if score > best_score:
            best_score = score
            best_pair = (city_col, cinema_col)
    return best_pair


def parse_cinema_report_xlsx(report_import):
    ensure_file_hash(report_import)
    workbook = load_workbook(report_import.source_file.path, data_only=True, read_only=False)
    created = 0
    report_import.rows.exclude(status=ImportStatus.IMPORTED).delete()

    workbook_text = " ".join(workbook.sheetnames)
    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            workbook_text += " " + " ".join(_xlsx_cell_text(value) for value in row if value)
    fallback_title = find_known_title_from_import(report_import, workbook_text)

    for ws in workbook.worksheets:
        metric_columns = _find_metric_columns(ws)
        if not metric_columns:
            continue
        city_col, cinema_col = _find_identity_columns(ws, metric_columns[0]["row"])
        first_data_row = metric_columns[0]["row"] + 1
        for row_index in range(first_data_row, ws.max_row + 1):
            city = _xlsx_cell_text(ws.cell(row_index, city_col).value)
            cinema_name = _xlsx_cell_text(ws.cell(row_index, cinema_col).value)
            if not city and not cinema_name:
                continue
            cinema = get_or_create_cinema(cinema_name)
            for metric in metric_columns:
                admissions = parse_int(ws.cell(row_index, metric["admissions_col"]).value)
                gross = parse_decimal(ws.cell(row_index, metric["gross_col"]).value)
                reported_rental = (
                    parse_decimal(ws.cell(row_index, metric["rental_col"]).value)
                    if metric["rental_col"]
                    else None
                )
                if admissions <= 0 and gross <= 0:
                    continue
                confidence = Decimal("40.00")
                if fallback_title:
                    confidence += Decimal("25.00")
                if cinema:
                    confidence += Decimal("20.00")
                if metric["date_from"]:
                    confidence += Decimal("15.00")
                identity = f"{ws.title}|{row_index}|{metric['admissions_col']}|{metric['date_label']}"
                upsert_import_row(
                    report_import,
                    identity,
                    {
                        "status": ImportStatus.NEEDS_REVIEW,
                        "title": fallback_title,
                        "cinema": cinema,
                        "city": city,
                        "date_from": metric["date_from"],
                        "date_to": metric["date_to"] or metric["date_from"],
                        "screenings": 0,
                        "admissions": admissions,
                        "box_office_gross": gross,
                        "reported_rental_amount": reported_rental,
                        "confidence": min(confidence, Decimal("100.00")),
                        "source_line": f"{ws.title} wiersz {row_index}, {metric['date_label']}: {city} / {cinema_name}",
                        "raw_payload": {
                            "sheet": ws.title,
                            "row": row_index,
                            "date_label": metric["date_label"],
                            "admissions_col": metric["admissions_col"],
                            "gross_col": metric["gross_col"],
                            "rental_col": metric["rental_col"],
                        },
                    },
                )
                created += 1

    if created == 0:
        upsert_import_row(
            report_import,
            "xlsx-fallback",
            {
                "status": ImportStatus.NEEDS_REVIEW,
                "title": fallback_title,
                "confidence": Decimal("10.00"),
                "source_line": "Nie znaleziono dodatnich wartości w kolumnach Widzowie.",
                "raw_payload": {"sheets": workbook.sheetnames},
                "notes": "Sprawdź, czy arkusz ma nagłówek Widzowie oraz dane widowni.",
            },
        )
        created = 1

    report_import.original_filename = report_import.original_filename or Path(report_import.source_file.name).name
    report_import.status = ImportStatus.NEEDS_REVIEW
    report_import.parsed_at = timezone.now()
    duplicate = CinemaReportImport.objects.filter(
        file_hash=report_import.file_hash,
        status=ImportStatus.IMPORTED,
    ).exclude(pk=report_import.pk).first()
    duplicate_note = f" Ten sam plik był już zaimportowany jako #{duplicate.pk}." if duplicate else ""
    report_import.parser_notes = f"Rozpoznano {created} wierszy z XLSX. Sprawdź dane przed akceptacją.{duplicate_note}"
    report_import.save(update_fields=["original_filename", "status", "parsed_at", "parser_notes", "updated_at"])
    return created


def parse_cinema_report_import(report_import):
    filename = report_import.source_file.name.lower()
    if filename.endswith(".pdf"):
        return parse_cinema_report_pdf(report_import)
    if filename.endswith(".xlsx"):
        return parse_cinema_report_xlsx(report_import)
    raise ValueError("Obslugiwane sa pliki PDF oraz XLSX.")


def prepare_import_for_week(report_import, week):
    report_import.target_booking_week = week
    report_import.save(update_fields=["target_booking_week", "updated_at"])
    for row in report_import.rows.exclude(status=ImportStatus.IMPORTED):
        update_fields = []
        if not row.title_id:
            row.title = week.booking.title
            update_fields.append("title")
        if not row.cinema_id:
            row.cinema = week.booking.cinema
            update_fields.append("cinema")
        if not row.date_from:
            row.date_from = week.date_from
            update_fields.append("date_from")
        if not row.date_to:
            row.date_to = week.date_to
            update_fields.append("date_to")
        if not row.currency:
            row.currency = week.currency
            update_fields.append("currency")
        if update_fields:
            row.save(update_fields=[*update_fields, "updated_at"])


def _candidate_week_for_row(row, preferred_week=None):
    if preferred_week:
        return preferred_week
    if row.booking_week_id:
        return row.booking_week
    if not all([row.title_id, row.cinema_id, row.date_from, row.date_to]):
        return None
    return CinemaBookingWeek.objects.select_related(
        "booking__title",
        "booking__cinema",
        "booking__crm_deal",
    ).filter(
        booking__title_id=row.title_id,
        booking__cinema_id=row.cinema_id,
        date_from__lte=row.date_to,
        date_to__gte=row.date_from,
    ).order_by("-date_from", "pk").first()


def review_import_row(row, preferred_week=None):
    week = _candidate_week_for_row(row, preferred_week or row.report_import.target_booking_week)
    errors = []
    warnings = []
    duplicate = None
    if row.source_fingerprint:
        duplicate = CinemaReportImportRow.objects.filter(
            source_fingerprint=row.source_fingerprint,
            status=ImportStatus.IMPORTED,
            booking_week__isnull=False,
        ).exclude(pk=row.pk).select_related("booking_week").first()

    if not row.title_id:
        errors.append("Nie rozpoznano tytułu.")
    if not row.cinema_id:
        errors.append("Nie rozpoznano kina.")
    if not row.date_from or not row.date_to:
        errors.append("Brakuje okresu raportu.")
    if not week:
        errors.append("Nie znaleziono pasującego tygodnia bookingu.")
    else:
        if row.title_id and row.title_id != week.booking.title_id:
            errors.append("Tytuł z raportu nie odpowiada wybranemu bookingowi.")
        if row.cinema_id and row.cinema_id != week.booking.cinema_id:
            errors.append("Kino z raportu nie odpowiada wybranemu bookingowi.")
        if row.date_from and row.date_to:
            if row.date_to < week.date_from or row.date_from > week.date_to:
                errors.append("Okres raportu nie pokrywa się z tygodniem grania.")
            elif row.date_from != week.date_from or row.date_to != week.date_to:
                warnings.append("Okres raportu częściowo pokrywa się z tygodniem grania.")
    if duplicate:
        errors.append(f"Ten sam wiersz został już zatwierdzony w imporcie #{duplicate.report_import_id}.")
    if row.confidence < Decimal("70.00"):
        warnings.append("Pewność automatycznego rozpoznania jest niższa niż 70%.")

    calculation = None
    values = None
    variance = None
    if week:
        values = week.calculation_values()
        calculation = week.preview_calculation(
            box_office_gross=row.box_office_gross,
            screenings=row.screenings,
        )
        if row.reported_rental_amount is None:
            warnings.append("Raport nie zawiera kwoty film rental; FILMERP obliczy ją z warunków bookingu.")
        else:
            variance = (row.reported_rental_amount - calculation.rental_amount).quantize(Decimal("0.01"))
            if abs(variance) > Decimal("0.01"):
                warnings.append("Kwota podana przez kino różni się od kalkulacji FILMERP.")

    return {
        "row": row,
        "week": week,
        "ready": not errors,
        "errors": errors,
        "warnings": warnings,
        "duplicate": duplicate,
        "calculation": calculation,
        "values": values,
        "variance": variance,
    }


def approve_import_rows(rows, *, approved_by=None):
    imported = 0
    skipped = 0
    for row in rows:
        if row.status == ImportStatus.IMPORTED:
            skipped += 1
            continue
        if not row.can_approve():
            skipped += 1
            continue
        row.approve(approved_by=approved_by)
        imported += 1
    imports = CinemaReportImport.objects.filter(rows__in=rows).distinct()
    for report_import in imports:
        if report_import.rows.exclude(status=ImportStatus.IMPORTED).exists():
            report_import.status = ImportStatus.NEEDS_REVIEW
        else:
            report_import.status = ImportStatus.IMPORTED
            report_import.imported_at = timezone.now()
        report_import.save(update_fields=["status", "imported_at", "updated_at"])
    return imported, skipped
