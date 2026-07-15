from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from distribution.sales_import import EXPECTED_HEADERS, import_sales_report_rows


class Command(BaseCommand):
    help = "Importuje raporty sprzedazy z XLSX. Pierwszy wiersz musi zawierac: " + ", ".join(EXPECTED_HEADERS)

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path")
        parser.add_argument("--sheet", help="Nazwa arkusza. Domyslnie aktywny arkusz.")
        parser.add_argument("--create-missing", action="store_true", help="Tworz brakujace tytuly/kontrahentow/terytoria.")

    def handle(self, *args, **options):
        workbook = load_workbook(options["xlsx_path"], data_only=True)
        worksheet = workbook[options["sheet"]] if options["sheet"] else workbook.active
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        missing_headers = [header for header in EXPECTED_HEADERS if header not in headers]
        if missing_headers:
            raise CommandError(f"Brakuje kolumn: {', '.join(missing_headers)}")

        rows = []
        for line_no, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row = dict(zip(headers, values))
            if any(row.get(header) not in (None, "") for header in EXPECTED_HEADERS):
                rows.append((line_no, row))
        created, updated = import_sales_report_rows(rows, create_missing=options["create_missing"])
        self.stdout.write(self.style.SUCCESS(f"Import zakonczony. Utworzono: {created}, zaktualizowano: {updated}."))
