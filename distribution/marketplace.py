import csv
from io import BytesIO, StringIO

from django.core.management.base import CommandError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import MarketplaceAgeRating, MarketplaceCondition, MarketplaceMediaType, Title


MARKETPLACE_HEADERS = [
    "title_pl",
    "original_title",
    "ean",
    "marketplace_category_id",
    "marketplace_category_name",
    "media_type",
    "condition",
    "genre",
    "director",
    "cast",
    "production_year",
    "runtime_minutes",
    "countries",
    "release_edition",
    "package_type",
    "discs_count",
    "region_code",
    "audio_languages",
    "subtitle_languages",
    "dubbing_languages",
    "lector_languages",
    "age_rating",
    "color_mode",
    "aspect_ratio",
    "description",
    "tags",
    "mg_advance",
    "acquisition_currency",
]

HEADER_TO_FIELD = {
    "condition": "marketplace_condition",
    "description": "marketplace_description",
    "tags": "marketplace_tags",
}

MEDIA_ALIASES = {
    "dvd": MarketplaceMediaType.DVD,
    "blu-ray": MarketplaceMediaType.BLURAY,
    "bluray": MarketplaceMediaType.BLURAY,
    "blu_ray": MarketplaceMediaType.BLURAY,
    "4k": MarketplaceMediaType.UHD_BLURAY,
    "4k uhd": MarketplaceMediaType.UHD_BLURAY,
    "uhd": MarketplaceMediaType.UHD_BLURAY,
    "digital": MarketplaceMediaType.DIGITAL,
    "vhs": MarketplaceMediaType.VHS,
}

CONDITION_ALIASES = {
    "nowy": MarketplaceCondition.NEW,
    "new": MarketplaceCondition.NEW,
    "uzywany": MarketplaceCondition.USED,
    "używany": MarketplaceCondition.USED,
    "used": MarketplaceCondition.USED,
    "odnowiony": MarketplaceCondition.REFURBISHED,
}

AGE_ALIASES = {
    "bez ograniczen": MarketplaceAgeRating.ALL,
    "bez ograniczeń": MarketplaceAgeRating.ALL,
    "all": MarketplaceAgeRating.ALL,
    "0": MarketplaceAgeRating.ALL,
    "7": MarketplaceAgeRating.AGE_7,
    "7+": MarketplaceAgeRating.AGE_7,
    "12": MarketplaceAgeRating.AGE_12,
    "12+": MarketplaceAgeRating.AGE_12,
    "15": MarketplaceAgeRating.AGE_15,
    "15+": MarketplaceAgeRating.AGE_15,
    "16": MarketplaceAgeRating.AGE_16,
    "16+": MarketplaceAgeRating.AGE_16,
    "18": MarketplaceAgeRating.AGE_18,
    "18+": MarketplaceAgeRating.AGE_18,
}


def normalize_choice(value, aliases):
    key = str(value or "").strip().lower()
    return aliases.get(key, value or "")


def title_to_marketplace_row(title):
    return {
        "title_pl": title.title_pl,
        "original_title": title.original_title,
        "ean": title.ean,
        "marketplace_category_id": title.marketplace_category_id,
        "marketplace_category_name": title.marketplace_category_name,
        "media_type": title.get_media_type_display() if title.media_type else "",
        "condition": title.get_marketplace_condition_display() if title.marketplace_condition else "",
        "genre": title.genre,
        "director": title.director,
        "cast": title.cast,
        "production_year": title.production_year or "",
        "runtime_minutes": title.runtime_minutes or "",
        "countries": title.countries,
        "release_edition": title.release_edition,
        "package_type": title.package_type,
        "discs_count": title.discs_count or "",
        "region_code": title.region_code,
        "audio_languages": title.audio_languages,
        "subtitle_languages": title.subtitle_languages,
        "dubbing_languages": title.dubbing_languages,
        "lector_languages": title.lector_languages,
        "age_rating": title.get_age_rating_display() if title.age_rating else "",
        "color_mode": title.color_mode,
        "aspect_ratio": title.aspect_ratio,
        "description": title.marketplace_description,
        "tags": title.marketplace_tags,
        "mg_advance": title.mg_advance,
        "acquisition_currency": title.acquisition_currency,
    }


def export_marketplace_csv(queryset):
    output = StringIO()
    output.write("\ufeff")
    writer = csv.DictWriter(output, fieldnames=MARKETPLACE_HEADERS)
    writer.writeheader()
    for title in queryset.order_by("title_pl"):
        writer.writerow(title_to_marketplace_row(title))
    return output.getvalue()


def export_marketplace_xlsx(queryset):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Marketplace catalog"
    sheet.append(MARKETPLACE_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF7")
    for title in queryset.order_by("title_pl"):
        row = title_to_marketplace_row(title)
        sheet.append([row[header] for header in MARKETPLACE_HEADERS])
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 45)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def import_marketplace_rows(rows, create_missing=False):
    created = 0
    updated = 0
    for line_no, row in rows:
        normalized = {str(key or "").strip(): value for key, value in row.items()}
        title_name = str(normalized.get("title_pl") or "").strip()
        if not title_name:
            raise CommandError(f"Linia {line_no}: brak title_pl.")
        title, was_created = Title.objects.get_or_create(title_pl=title_name) if create_missing else (None, False)
        if not create_missing:
            try:
                title = Title.objects.get(title_pl=title_name)
            except Title.DoesNotExist as exc:
                raise CommandError(f"Linia {line_no}: brak tytulu: {title_name}") from exc
        for header in MARKETPLACE_HEADERS:
            if header == "title_pl" or header not in normalized:
                continue
            field = HEADER_TO_FIELD.get(header, header)
            value = normalized.get(header)
            if value is None:
                continue
            if header == "media_type":
                value = normalize_choice(value, MEDIA_ALIASES)
            elif header == "condition":
                value = normalize_choice(value, CONDITION_ALIASES)
            elif header == "age_rating":
                value = normalize_choice(value, AGE_ALIASES)
            elif header in {"production_year", "runtime_minutes", "discs_count"}:
                value = int(value) if str(value).strip() else None
            setattr(title, field, value)
        title.save()
        if was_created:
            created += 1
        else:
            updated += 1
    return created, updated


def load_csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        missing = [header for header in ["title_pl"] if header not in (reader.fieldnames or [])]
        if missing:
            raise CommandError(f"Brakuje kolumn: {', '.join(missing)}")
        return list(enumerate(reader, start=2))


def load_xlsx_rows(path, sheet_name=None):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if "title_pl" not in headers:
        raise CommandError("Brakuje kolumny: title_pl")
    rows = []
    for line_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in row.values()):
            rows.append((line_no, row))
    return rows
