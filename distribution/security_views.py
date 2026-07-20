from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import Group, Permission
from django.contrib.auth.tokens import default_token_generator
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from allauth.account.internal.flows.password_reset import request_password_reset
from allauth.mfa.models import Authenticator
from allauth.usersessions.models import UserSession
from auditlog.models import LogEntry
from axes.utils import reset as reset_axes

from .models import AuditAction, AuditEvent, LoginEvent, LoginEventType, SecurityProfile
from .security import describe_user_agent, record_audit_event, record_login_event
from .security_forms import DeactivateUserForm, SYSTEM_ROLE_NAMES, UserAccountForm, UserCreateForm


User = get_user_model()


def _safe_cell(value):
    value = "" if value is None else str(value)
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _tabular_export(filename, headers, rows, export_format):
    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")
        writer.writerow(headers)
        writer.writerows([[_safe_cell(value) for value in row] for row in rows])
        return response

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FILMERP"
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0058F8")
    for row in rows:
        worksheet.append([_safe_cell(value) for value in row])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def _send_password_link(request, user, invitation=False):
    if not user.email:
        raise ValueError("Konto nie ma adresu e-mail.")
    request_password_reset(request, user.email, [user], default_token_generator)
    profile, _ = SecurityProfile.objects.get_or_create(user=user)
    now = timezone.now()
    if invitation and not profile.invited_at:
        profile.invited_at = now
        profile.invited_by = request.user
    profile.invitation_sent_at = now
    profile.save(update_fields=["invited_at", "invited_by", "invitation_sent_at", "updated_at"])
    return "console.EmailBackend" not in settings.EMAIL_BACKEND


@login_required
@permission_required("distribution.manage_users", raise_exception=True)
def account_list(request):
    users = User.objects.select_related("security_profile").prefetch_related("groups").order_by("username")
    query = request.GET.get("q", "").strip()
    role = request.GET.get("role", "").strip()
    status = request.GET.get("status", "").strip()
    mfa = request.GET.get("mfa", "").strip()
    if query:
        users = users.filter(
            Q(username__icontains=query)
            | Q(email__icontains=query)
            | Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
        )
    if role:
        users = users.filter(groups__name=role)
    if status == "active":
        users = users.filter(is_active=True).exclude(password__startswith="!")
    elif status == "inactive":
        users = users.filter(is_active=False)
    elif status == "invited":
        users = users.filter(is_active=True, password__startswith="!")

    user_ids = list(users.values_list("pk", flat=True))
    mfa_ids = set(
        Authenticator.objects.filter(user_id__in=user_ids, type=Authenticator.Type.TOTP).values_list("user_id", flat=True)
    )
    if mfa == "enabled":
        users = users.filter(pk__in=mfa_ids)
    elif mfa == "disabled":
        users = users.exclude(pk__in=mfa_ids)
    session_counts = Counter(
        UserSession.objects.filter(user_id__in=user_ids).values_list("user_id", flat=True)
    )
    rows = [
        {
            "user": user,
            "roles": list(user.groups.filter(name__in=SYSTEM_ROLE_NAMES).order_by("name")),
            "mfa_enabled": user.pk in mfa_ids,
            "mfa_required": user.security_profile.mfa_required,
            "is_invited": not user.has_usable_password(),
            "session_count": session_counts[user.pk],
        }
        for user in users.distinct()
    ]
    page = Paginator(rows, 50).get_page(request.GET.get("page"))
    return render(
        request,
        "security/account_list.html",
        {
            "page": page,
            "roles": Group.objects.filter(name__in=SYSTEM_ROLE_NAMES).order_by("name"),
            "query": query,
            "selected_role": role,
            "selected_status": status,
            "selected_mfa": mfa,
        },
    )


@login_required
@permission_required("distribution.manage_users", raise_exception=True)
def account_create(request):
    form = UserCreateForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        delivered = None
        if form.cleaned_data["send_invitation"]:
            delivered = _send_password_link(request, user, invitation=True)
        record_audit_event(
            AuditAction.CREATE,
            f"Utworzono konto {user.username}.",
            request=request,
            module="users",
            instance=user,
            changes={"roles": [group.name for group in user.groups.all()]},
        )
        if delivered is True:
            messages.success(request, "Konto zostało utworzone. Zaproszenie wysłano e-mailem.")
        elif delivered is False:
            messages.warning(request, "Konto zostało utworzone, ale serwer poczty nie jest skonfigurowany. Link zapisano w logu aplikacji.")
        else:
            messages.success(request, "Konto zostało utworzone.")
        return redirect("security:account_detail", pk=user.pk)
    return render(request, "security/account_form.html", {"form": form, "form_title": "Dodaj uzytkownika"})


@login_required
@permission_required("distribution.manage_users", raise_exception=True)
def account_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    before_roles = list(user.groups.filter(name__in=SYSTEM_ROLE_NAMES).values_list("name", flat=True))
    form = UserAccountForm(request.POST or None, instance=user)
    if request.method == "POST" and form.is_valid():
        if user == request.user and not form.cleaned_data["is_active"]:
            form.add_error("is_active", "Nie mozesz dezaktywowac aktualnie uzywanego konta.")
        else:
            user = form.save()
            after_roles = list(user.groups.filter(name__in=SYSTEM_ROLE_NAMES).values_list("name", flat=True))
            record_audit_event(
                AuditAction.UPDATE,
                f"Zmieniono konto {user.username}.",
                request=request,
                module="users",
                instance=user,
                changes={"roles": {"old": before_roles, "new": after_roles}, "mfa_required": form.cleaned_data["mfa_required"], "force_password_change": form.cleaned_data["force_password_change"]},
            )
            messages.success(request, "Zmiany konta zostaly zapisane.")
            return redirect("security:account_detail", pk=user.pk)
    return render(request, "security/account_form.html", {"form": form, "form_title": f"Edytuj konto: {user.username}", "account": user})


@login_required
@permission_required("distribution.manage_users", raise_exception=True)
def account_detail(request, pk):
    account = get_object_or_404(User.objects.select_related("security_profile").prefetch_related("groups"), pk=pk)
    sessions = []
    active_sessions = UserSession.objects.purge_and_list(account)
    for session in sorted(active_sessions, key=lambda item: item.last_seen_at, reverse=True):
        sessions.append({"session": session, **describe_user_agent(session.user_agent)})
    return render(
        request,
        "security/account_detail.html",
        {
            "account": account,
            "sessions": sessions,
            "mfa_authenticators": Authenticator.objects.filter(user=account).order_by("type", "-created_at"),
            "login_events": LoginEvent.objects.filter(user=account)[:12],
            "audit_events": AuditEvent.objects.filter(actor=account)[:12],
            "deactivate_form": DeactivateUserForm(),
        },
    )


@login_required
@permission_required("distribution.manage_users", raise_exception=True)
@require_POST
def account_action(request, pk):
    account = get_object_or_404(User, pk=pk)
    action = request.POST.get("action", "")
    profile, _ = SecurityProfile.objects.get_or_create(user=account)

    if action in {"invite", "reset_password"}:
        try:
            delivered = _send_password_link(request, account, invitation=action == "invite")
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("security:account_detail", pk=account.pk)
        record_audit_event(
            AuditAction.SECURITY,
            f"Wyslano link ustawienia hasla dla {account.username}.",
            request=request,
            module="security",
            instance=account,
        )
        if delivered:
            messages.success(request, "Link do ustawienia hasła został wysłany.")
        else:
            messages.warning(request, "Serwer poczty nie jest skonfigurowany. Link zapisano w logu aplikacji.")
    elif action == "deactivate":
        form = DeactivateUserForm(request.POST)
        if account == request.user:
            messages.error(request, "Nie mozesz dezaktywowac aktualnie uzywanego konta.")
        elif account.is_superuser:
            messages.error(request, "Konta superadministratora nie mozna dezaktywowac z tego ekranu.")
        elif form.is_valid():
            account.is_active = False
            account.save(update_fields=["is_active"])
            profile.deactivated_at = timezone.now()
            profile.deactivated_by = request.user
            profile.deactivation_reason = form.cleaned_data["reason"]
            profile.save(update_fields=["deactivated_at", "deactivated_by", "deactivation_reason", "updated_at"])
            for session in list(UserSession.objects.filter(user=account)):
                session.end()
            record_audit_event(
                AuditAction.SECURITY,
                f"Dezaktywowano konto {account.username}.",
                request=request,
                module="users",
                instance=account,
                metadata={"reason": form.cleaned_data["reason"]},
            )
            messages.success(request, "Konto zostalo dezaktywowane, a jego sesje zakonczone.")
    elif action == "reactivate":
        account.is_active = True
        account.save(update_fields=["is_active"])
        profile.deactivated_at = None
        profile.deactivated_by = None
        profile.deactivation_reason = ""
        profile.save(update_fields=["deactivated_at", "deactivated_by", "deactivation_reason", "updated_at"])
        record_audit_event(AuditAction.SECURITY, f"Reaktywowano konto {account.username}.", request=request, module="users", instance=account)
        messages.success(request, "Konto zostalo ponownie aktywowane.")
    elif action == "revoke_sessions":
        ended = 0
        for session in list(UserSession.objects.filter(user=account)):
            session.end()
            ended += 1
        record_audit_event(
            AuditAction.FORCE_LOGOUT,
            f"Zakonczono sesje konta {account.username}.",
            request=request,
            module="sessions",
            instance=account,
            metadata={"ended_sessions": ended},
        )
        record_login_event(LoginEventType.LOGOUT_FORCED, request=request, user=account, reason=f"Wymuszone przez {request.user.username}")
        messages.success(request, f"Zakonczono aktywne sesje: {ended}.")
    elif action == "reset_mfa":
        removed = Authenticator.objects.filter(user=account).count()
        Authenticator.objects.filter(user=account).delete()
        for session in list(UserSession.objects.filter(user=account)):
            session.end()
        record_audit_event(
            AuditAction.SECURITY,
            f"Zresetowano MFA konta {account.username}.",
            request=request,
            module="security",
            instance=account,
            metadata={"removed_authenticators": removed},
        )
        record_login_event(LoginEventType.MFA_DISABLED, request=request, user=account, reason=f"Reset przez {request.user.username}")
        messages.success(request, "MFA zostalo zresetowane. Przy kolejnym logowaniu konto skonfiguruje je ponownie.")
    elif action == "unlock":
        removed = reset_axes(username=account.username)
        record_login_event(LoginEventType.UNLOCKED, request=request, user=account, reason=f"Odblokowane przez {request.user.username}")
        record_audit_event(AuditAction.SECURITY, f"Odblokowano logowanie konta {account.username}.", request=request, module="security", instance=account, metadata={"removed_attempts": removed})
        messages.success(request, "Usunieto blokade nieudanych prob logowania.")
    else:
        messages.error(request, "Nieznana operacja.")
    return redirect("security:account_detail", pk=account.pk)


@login_required
@permission_required("distribution.manage_users", raise_exception=True)
def role_list(request):
    roles = Group.objects.filter(name__in=SYSTEM_ROLE_NAMES).prefetch_related("permissions", "user_set").order_by("name")
    return render(request, "security/role_list.html", {"roles": roles})


@login_required
@permission_required("distribution.manage_users", raise_exception=True)
def role_detail(request, pk):
    role = get_object_or_404(Group, pk=pk, name__in=SYSTEM_ROLE_NAMES)
    allowed = Permission.objects.filter(
        content_type__app_label__in=["distribution", "auth", "auditlog", "usersessions", "mfa"]
    ).select_related("content_type").order_by("content_type__app_label", "content_type__model", "codename")
    if request.method == "POST":
        if role.name == "administrator":
            messages.error(request, "Rola administrator ma staly, pelny zakres uprawnien.")
            return redirect("security:role_detail", pk=role.pk)
        selected = allowed.filter(pk__in=request.POST.getlist("permissions"))
        before = set(role.permissions.values_list("codename", flat=True))
        role.permissions.set(selected)
        after = set(selected.values_list("codename", flat=True))
        record_audit_event(
            AuditAction.PERMISSION,
            f"Zmieniono uprawnienia roli {role.name}.",
            request=request,
            module="roles",
            instance=role,
            changes={"added": sorted(after - before), "removed": sorted(before - after)},
        )
        messages.success(request, "Uprawnienia roli zostaly zapisane.")
        return redirect("security:role_detail", pk=role.pk)

    selected_ids = set(role.permissions.values_list("pk", flat=True))
    groups = defaultdict(list)
    for permission in allowed:
        key = f"{permission.content_type.app_label}.{permission.content_type.model}"
        groups[key].append({"permission": permission, "selected": permission.pk in selected_ids})
    return render(request, "security/role_detail.html", {"role": role, "permission_groups": dict(groups)})


def _filter_login_events(request):
    events = LoginEvent.objects.select_related("user")
    user_id = request.GET.get("user", "")
    event_type = request.GET.get("event", "")
    result = request.GET.get("result", "")
    ip = request.GET.get("ip", "").strip()
    date_from = parse_date(request.GET.get("date_from", ""))
    date_to = parse_date(request.GET.get("date_to", ""))
    if user_id.isdigit():
        events = events.filter(user_id=user_id)
    if event_type:
        events = events.filter(event_type=event_type)
    if result:
        events = events.filter(result=result)
    if ip:
        events = events.filter(ip_address__icontains=ip)
    if date_from:
        events = events.filter(occurred_at__date__gte=date_from)
    if date_to:
        events = events.filter(occurred_at__date__lte=date_to)
    return events


@login_required
@permission_required("distribution.view_security_log", raise_exception=True)
def login_history(request):
    events = _filter_login_events(request)
    export_format = request.GET.get("format")
    if export_format in {"csv", "xlsx"}:
        if not request.user.has_perm("distribution.export_audit_log"):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        rows = [
            [event.occurred_at.isoformat(), event.user.username if event.user else event.identifier, event.get_event_type_display(), event.result_label, event.ip_address, event.browser, event.operating_system, event.device, event.reason, event.request_id]
            for event in events[:10000]
        ]
        record_audit_event(AuditAction.EXPORT, "Wyeksportowano historie logowan.", request=request, module="security", metadata={"format": export_format, "rows": len(rows)})
        return _tabular_export("filmerp-historia-logowan", ["Czas", "Uzytkownik", "Zdarzenie", "Wynik", "IP", "Przegladarka", "System", "Urzadzenie", "Powod", "Request ID"], rows, export_format)
    page = Paginator(events, 50).get_page(request.GET.get("page"))
    return render(request, "security/login_history.html", {"page": page, "users": User.objects.order_by("username"), "event_types": LoginEventType.choices})


def _audit_rows(request):
    actor_id = request.GET.get("user", "")
    action = request.GET.get("action", "")
    module = request.GET.get("module", "").strip()
    query = request.GET.get("q", "").strip()
    date_from = parse_date(request.GET.get("date_from", ""))
    date_to = parse_date(request.GET.get("date_to", ""))

    semantic = AuditEvent.objects.select_related("actor", "content_type")
    automatic = LogEntry.objects.select_related("actor", "content_type")
    if actor_id.isdigit():
        semantic = semantic.filter(actor_id=actor_id)
        automatic = automatic.filter(actor_id=actor_id)
    if action:
        semantic = semantic.filter(action=action)
        automatic_action = {"create": LogEntry.Action.CREATE, "update": LogEntry.Action.UPDATE, "delete": LogEntry.Action.DELETE}.get(action)
        automatic = automatic.filter(action=automatic_action) if automatic_action is not None else automatic.none()
    if module:
        semantic = semantic.filter(module=module)
        automatic = automatic.filter(content_type__model__icontains=module)
    if query:
        semantic = semantic.filter(Q(summary__icontains=query) | Q(object_repr__icontains=query) | Q(object_pk__icontains=query))
        automatic = automatic.filter(Q(object_repr__icontains=query) | Q(object_pk__icontains=query))
    if date_from:
        semantic = semantic.filter(occurred_at__date__gte=date_from)
        automatic = automatic.filter(timestamp__date__gte=date_from)
    if date_to:
        semantic = semantic.filter(occurred_at__date__lte=date_to)
        automatic = automatic.filter(timestamp__date__lte=date_to)

    rows = []
    for event in semantic[:5000]:
        rows.append({
            "timestamp": event.occurred_at,
            "actor": event.actor,
            "action": event.get_action_display(),
            "action_code": event.action,
            "module": event.module,
            "object": event.object_repr,
            "summary": event.summary,
            "changes": event.changes,
            "source": event.source,
            "request_id": event.request_id,
            "ip": event.ip_address,
            "integrity_ok": event.verify_integrity(),
        })
    automatic_labels = {
        LogEntry.Action.CREATE: "Utworzenie",
        LogEntry.Action.UPDATE: "Zmiana",
        LogEntry.Action.DELETE: "Usunięcie",
        LogEntry.Action.ACCESS: "Dostęp",
    }
    for event in automatic[:5000]:
        rows.append({
            "timestamp": event.timestamp,
            "actor": event.actor,
            "action": automatic_labels.get(event.action, "Zdarzenie"),
            "action_code": "automatic",
            "module": event.content_type.model if event.content_type else "system",
            "object": event.object_repr,
            "summary": "Automatyczny zapis zmiany rekordu.",
            "changes": event.changes_dict,
            "source": "admin" if event.remote_addr else "system",
            "request_id": event.cid,
            "ip": event.remote_addr,
            "integrity_ok": None,
        })
    rows.sort(key=lambda row: row["timestamp"], reverse=True)
    return rows


@login_required
@permission_required("distribution.view_business_audit", raise_exception=True)
def audit_history(request):
    rows = _audit_rows(request)
    export_format = request.GET.get("format")
    if export_format in {"csv", "xlsx"}:
        if not request.user.has_perm("distribution.export_audit_log"):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        export_rows = [
            [row["timestamp"].isoformat(), row["actor"] or "System", row["action"], row["module"], row["object"], row["summary"], row["changes"], row["source"], row["ip"], row["request_id"]]
            for row in rows[:10000]
        ]
        record_audit_event(AuditAction.EXPORT, "Wyeksportowano historie zmian.", request=request, module="security", metadata={"format": export_format, "rows": len(export_rows)})
        return _tabular_export("filmerp-historia-zmian", ["Czas", "Uzytkownik", "Akcja", "Modul", "Obiekt", "Opis", "Zmiany", "Zrodlo", "IP", "Request ID"], export_rows, export_format)
    page = Paginator(rows, 50).get_page(request.GET.get("page"))
    modules = sorted({row["module"] for row in rows})
    return render(request, "security/audit_history.html", {"page": page, "users": User.objects.order_by("username"), "actions": AuditAction.choices, "modules": modules})


@login_required
def mfa_required(request):
    enabled = Authenticator.objects.filter(user=request.user, type=Authenticator.Type.TOTP).exists()
    if enabled:
        return redirect("distribution:dashboard")
    return render(request, "security/mfa_required.html")
