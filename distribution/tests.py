from datetime import date
from decimal import Decimal
from io import BytesIO
import shutil
import tempfile
from zipfile import ZipFile

from django.contrib import admin
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.db.migrations.executor import MigrationExecutor
from django.test import TestCase, TransactionTestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .models import (
    Cost,
    CostCategory,
    CostScopeMode,
    AcquisitionAgreement,
    Counterparty,
    Currency,
    DocumentInboxItem,
    DocumentStatus,
    DocumentType,
    ExploitationField,
    RoyaltyStatement,
    SalesReport,
    WaterfallAllocationMode,
    WaterfallPlan,
    WaterfallParticipant,
    WaterfallRecoupmentItem,
    WaterfallRecoupmentRule,
    WaterfallRun,
    WaterfallRunCostAllocation,
    WaterfallRunStatus,
    WaterfallStep,
    WaterfallStepType,
    Title,
)
from .documents import classify_document
from .roles import sync_role_groups
from .waterfall_engine import calculate_waterfall_run, finalize_waterfall_run


class LegacyWaterfallMigrationTests(TransactionTestCase):
    reset_sequences = True

    def test_legacy_rule_is_migrated_when_title_has_no_current_plan(self):
        executor = MigrationExecutor(connection)
        executor.migrate([("distribution", "0011_documentinboxitem")])
        old_apps = executor.loader.project_state([("distribution", "0011_documentinboxitem")]).apps
        CounterpartyOld = old_apps.get_model("distribution", "Counterparty")
        TitleOld = old_apps.get_model("distribution", "Title")
        LegacyRule = old_apps.get_model("distribution", "WaterfallRecoupmentRule")
        producer = CounterpartyOld.objects.create(name="Producent migracji")
        title = TitleOld.objects.create(title_pl="Film starego waterfallu", producer_id=producer.pk)
        LegacyRule.objects.create(
            title_id=title.pk,
            exploitation_field="cinema",
            currency="PLN",
            recoupment_pool=Decimal("50000.00"),
            distributor_fee_percent=Decimal("25.00"),
            participant_share_percent=Decimal("50.00"),
            include_recoupable_costs=True,
            fee_after_recoupment=True,
            active=True,
        )

        executor = MigrationExecutor(connection)
        executor.migrate([("distribution", "0012_alter_waterfallplan_options_and_more")])
        new_apps = executor.loader.project_state([("distribution", "0012_alter_waterfallplan_options_and_more")]).apps
        Plan = new_apps.get_model("distribution", "WaterfallPlan")
        Step = new_apps.get_model("distribution", "WaterfallStep")
        plan = Plan.objects.get(title_id=title.pk)
        self.assertEqual(plan.exploitation_fields, ["cinema"])
        self.assertEqual(plan.status, "active")
        self.assertEqual(list(Step.objects.filter(plan_id=plan.pk).order_by("phase", "sort_order").values_list("name", flat=True)), [
            "Bazowy recoupment i koszty",
            "Fee dystrybutora",
            "Udział producenta / licencjodawcy",
        ])


class WaterfallV2Tests(TestCase):
    def setUp(self):
        self.title = Title.objects.create(title_pl="Film testowy", mg_advance=Decimal("50.00"))
        self.distributor = Counterparty.objects.create(name="Dystrybutor")
        self.producer = Counterparty.objects.create(name="Producent")
        self.investor = Counterparty.objects.create(name="Inwestor")
        self.cinema = Counterparty.objects.create(name="Kino")
        self.plan = WaterfallPlan.objects.create(title=self.title, name="Plan PISF", currency=Currency.PLN)

    def add_sales(self, start, end, gross="1000.00", field=ExploitationField.CINEMA, currency=Currency.PLN):
        return SalesReport.objects.create(
            title=self.title,
            counterparty=self.cinema,
            exploitation_field=field,
            period_start=start,
            period_end=end,
            currency=currency,
            gross_revenue=Decimal(gross),
        )

    def test_pisf_style_phases_and_parallel_hard_money(self):
        self.add_sales(date(2026, 1, 1), date(2026, 3, 31))
        WaterfallStep.objects.create(
            plan=self.plan, phase=0, sort_order=10, name="Fee dystrybutora",
            step_type=WaterfallStepType.COMMISSION, beneficiary=self.distributor, percentage=Decimal("10.00"),
        )
        WaterfallStep.objects.create(
            plan=self.plan, phase=0, sort_order=20, name="P&A",
            step_type=WaterfallStepType.RECOUPMENT, beneficiary=self.distributor, target_amount=Decimal("200.00"),
        )
        for beneficiary in (self.producer, self.investor):
            WaterfallStep.objects.create(
                plan=self.plan, phase=1, name=f"Hard money {beneficiary.name}",
                step_type=WaterfallStepType.RECOUPMENT,
                allocation_mode=WaterfallAllocationMode.PARI_PASSU,
                beneficiary=beneficiary,
                target_amount=Decimal("300.00"),
            )

        run = calculate_waterfall_run(WaterfallRun.objects.create(
            plan=self.plan, period_start=date(2026, 1, 1), period_end=date(2026, 3, 31),
        ))

        self.assertEqual(run.net_revenue, Decimal("1000.00"))
        self.assertEqual(list(run.lines.values_list("allocated_amount", flat=True)), [
            Decimal("100.00"), Decimal("200.00"), Decimal("300.00"), Decimal("300.00"),
        ])
        self.assertEqual(run.closing_available, Decimal("100.00"))

    def test_finalized_recoupment_carries_forward(self):
        step = WaterfallStep.objects.create(
            plan=self.plan, phase=0, name="MG", step_type=WaterfallStepType.RECOUPMENT,
            beneficiary=self.producer, target_amount=Decimal("500.00"),
        )
        self.add_sales(date(2026, 1, 1), date(2026, 3, 31), gross="300.00")
        first = calculate_waterfall_run(WaterfallRun.objects.create(
            plan=self.plan, period_start=date(2026, 1, 1), period_end=date(2026, 3, 31),
        ))
        finalize_waterfall_run(first)
        self.add_sales(date(2026, 4, 1), date(2026, 6, 30), gross="300.00")
        second = calculate_waterfall_run(WaterfallRun.objects.create(
            plan=self.plan, period_start=date(2026, 4, 1), period_end=date(2026, 6, 30),
        ))

        line = second.lines.get(step=step)
        self.assertEqual(line.opening_recoupment, Decimal("200.00"))
        self.assertEqual(line.allocated_amount, Decimal("200.00"))
        self.assertEqual(second.closing_available, Decimal("100.00"))

    def test_recoupable_cost_scope_and_statement_freeze(self):
        self.plan.applies_to_all_exploitation_fields = False
        self.plan.exploitation_fields = [ExploitationField.CINEMA]
        self.plan.save()
        self.add_sales(date(2026, 1, 1), date(2026, 3, 31), gross="1000.00")
        self.add_sales(date(2026, 1, 1), date(2026, 3, 31), gross="900.00", field=ExploitationField.LINEAR_TV)
        Cost.objects.create(
            title=self.title, category=CostCategory.PA, cost_date=date(2026, 2, 1),
            currency=Currency.PLN, net_amount=Decimal("100.00"), recoupable=True,
            scope_mode=CostScopeMode.SELECTED,
            scope_fields=[ExploitationField.CINEMA],
        )
        statement = RoyaltyStatement.objects.create(
            title=self.title, recipient=self.producer,
            period_start=date(2026, 1, 1), period_end=date(2026, 3, 31),
            currency=Currency.PLN, waterfall_plan=self.plan,
            distributor_fee_percent=Decimal("10.00"), recipient_share_percent=Decimal("50.00"),
        )
        statement.freeze_calculation(lock=True)
        self.assertEqual(statement.gross_revenue, Decimal("1000.00"))
        self.assertEqual(statement.amount_due, Decimal("400.00"))

        self.add_sales(date(2026, 2, 1), date(2026, 2, 28), gross="1000.00")
        statement.refresh_from_db()
        self.assertEqual(statement.gross_revenue, Decimal("1000.00"))
        self.assertEqual(statement.amount_due, Decimal("400.00"))

    def test_currencies_are_kept_separate_on_title(self):
        self.add_sales(date(2026, 1, 1), date(2026, 1, 31), gross="100.00", currency=Currency.PLN)
        self.add_sales(date(2026, 1, 1), date(2026, 1, 31), gross="50.00", currency=Currency.EUR)
        rows = {row["currency"]: row for row in self.title.financial_summary_by_currency}
        self.assertEqual(rows[Currency.PLN]["net"], Decimal("100.00"))
        self.assertEqual(rows[Currency.EUR]["net"], Decimal("50.00"))
        self.assertEqual(set(rows), {Currency.PLN, Currency.EUR})

    def test_pa_percentage_split_uses_field_pools_and_never_recoups_twice(self):
        cost = Cost.objects.create(
            title=self.title,
            category=CostCategory.PA,
            cost_date=date(2026, 1, 10),
            currency=Currency.PLN,
            net_amount=Decimal("500.00"),
            recoupable=True,
            scope_mode=CostScopeMode.ALLOCATED,
            allocation_percentages={
                ExploitationField.CINEMA: "80.00",
                ExploitationField.SVOD: "20.00",
            },
        )
        step = WaterfallStep.objects.create(
            plan=self.plan,
            phase=0,
            name="Zwrot P&A według pól",
            step_type=WaterfallStepType.RECOUPMENT,
            beneficiary=self.distributor,
            include_recoupable_costs=True,
        )
        self.add_sales(date(2026, 1, 1), date(2026, 3, 31), gross="100.00", field=ExploitationField.CINEMA)
        self.add_sales(date(2026, 1, 1), date(2026, 3, 31), gross="400.00", field=ExploitationField.SVOD)
        first = calculate_waterfall_run(WaterfallRun.objects.create(
            plan=self.plan, period_start=date(2026, 1, 1), period_end=date(2026, 3, 31),
        ))

        allocations = list(
            WaterfallRunCostAllocation.objects.filter(run_line__run=first, cost=cost)
            .order_by("exploitation_field")
            .values_list("exploitation_field", "allocated_amount")
        )
        self.assertEqual(allocations, [
            (ExploitationField.CINEMA, Decimal("100.00")),
            (ExploitationField.SVOD, Decimal("100.00")),
        ])
        self.assertEqual(first.lines.get(step=step).allocated_amount, Decimal("200.00"))
        finalize_waterfall_run(first)

        self.add_sales(date(2026, 4, 1), date(2026, 6, 30), gross="300.00", field=ExploitationField.CINEMA)
        second = calculate_waterfall_run(WaterfallRun.objects.create(
            plan=self.plan, period_start=date(2026, 4, 1), period_end=date(2026, 6, 30),
        ))
        self.assertEqual(second.lines.get(step=step).allocated_amount, Decimal("300.00"))
        finalize_waterfall_run(second)
        self.assertEqual(cost.recouped_amount, Decimal("500.00"))
        self.assertEqual(cost.outstanding_recoupment, Decimal("0.00"))
        statement = RoyaltyStatement.objects.create(
            title=self.title,
            recipient=self.distributor,
            period_start=date(2026, 4, 1),
            period_end=date(2026, 6, 30),
            currency=Currency.PLN,
            waterfall_plan=self.plan,
            waterfall_run=second,
        )
        statement.freeze_calculation(lock=True)
        self.assertEqual(statement.recoupable_costs, Decimal("300.00"))
        self.assertEqual(list(statement.recoupable_costs_queryset()), [cost])

        self.add_sales(date(2026, 7, 1), date(2026, 9, 30), gross="500.00", field=ExploitationField.CINEMA)
        third = calculate_waterfall_run(WaterfallRun.objects.create(
            plan=self.plan, period_start=date(2026, 7, 1), period_end=date(2026, 9, 30),
        ))
        self.assertEqual(third.lines.get(step=step).allocated_amount, Decimal("0.00"))
        self.assertFalse(WaterfallRunCostAllocation.objects.filter(run_line__run=third, cost=cost).exists())


class SettlementWorkflowTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.settings_override = self.settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)
        self.addCleanup(shutil.rmtree, self.media_root, True)

        self.user = get_user_model().objects.create_superuser(username="finance", email="", password="test-pass")
        self.client.force_login(self.user)
        self.title = Title.objects.create(title_pl="Miesieczny film")
        self.producer = Counterparty.objects.create(name="Producent statementu")
        self.cinema = Counterparty.objects.create(name="Kino raportujace")
        self.plan = WaterfallPlan.objects.create(
            title=self.title,
            name="Umowne rozliczenie",
            status="active",
            currency=Currency.PLN,
        )
        WaterfallStep.objects.create(
            plan=self.plan,
            phase=1,
            name="Udzial producenta",
            step_type=WaterfallStepType.SPLIT,
            beneficiary=self.producer,
            percentage=Decimal("50.00"),
        )
        SalesReport.objects.create(
            title=self.title,
            counterparty=self.cinema,
            exploitation_field=ExploitationField.CINEMA,
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            currency=Currency.PLN,
            gross_revenue=Decimal("1000.00"),
        )

    def workflow_payload(self):
        return {
            "title": str(self.title.pk),
            "plan": str(self.plan.pk),
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
            "currency": Currency.PLN,
        }

    def test_common_document_names_are_classified(self):
        self.assertEqual(classify_document("sample_cinema_report.pdf")[0], DocumentType.CINEMA_REPORT)
        self.assertEqual(classify_document("FV_2026_001.pdf")[0], DocumentType.COST_INVOICE)

    def test_readonly_role_can_view_but_cannot_upload_documents(self):
        sync_role_groups()
        readonly_user = get_user_model().objects.create_user(username="readonly-user", password="test-pass")
        readonly_user.groups.add(Group.objects.get(name="readonly"))
        self.client.force_login(readonly_user)

        response = self.client.get(reverse("distribution:document_center"))
        self.assertEqual(response.status_code, 200)
        response = self.client.post(reverse("distribution:document_center"), {
            "action": "upload",
            "source_file": SimpleUploadedFile("FV.pdf", b"invoice", content_type="application/pdf"),
        })
        self.assertEqual(response.status_code, 403)
        self.assertEqual(DocumentInboxItem.objects.count(), 0)
        response = self.client.get(reverse("distribution:contract_waterfall_wizard"))
        self.assertEqual(response.status_code, 403)
        response = self.client.get(reverse("distribution:title_create"))
        self.assertEqual(response.status_code, 403)

    def test_app_navigation_is_consistent(self):
        page_urls = [
            reverse("distribution:dashboard"),
            reverse("distribution:document_center"),
            reverse("distribution:title_list"),
            reverse("distribution:title_detail", args=[self.title.pk]),
            reverse("distribution:avails"),
            reverse("distribution:reports"),
            reverse("distribution:settlement_workbench"),
            reverse("distribution:statement_center"),
        ]
        expected_links = [
            reverse("distribution:dashboard"),
            reverse("distribution:title_list"),
            reverse("distribution:document_center"),
            reverse("distribution:avails"),
            reverse("distribution:reports"),
            reverse("distribution:settlement_workbench"),
            reverse("distribution:statement_center"),
            "/admin/",
        ]

        for page_url in page_urls:
            with self.subTest(page_url=page_url):
                response = self.client.get(page_url)
                self.assertEqual(response.status_code, 200)
                shell_html = response.content.decode().split('<aside class="app-sidebar"', 1)[1].split("</aside>", 1)[0]
                positions = [shell_html.index(f'href="{link}"') for link in expected_links]
                self.assertEqual(positions, sorted(positions))
                self.assertIn('id="sidebar-toggle"', shell_html)
                self.assertIn("lucide-sidebar", shell_html)
                self.assertContains(response, "distribution/filmerp-sidebar.js")
                self.assertContains(response, "distribution/filmerp-favicon.svg")

    def test_dashboard_and_title_card_use_title_centric_workflow(self):
        response = self.client.get(reverse("distribution:dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Twoje tytuły")
        self.assertContains(response, self.title.title_pl)
        for label in ("Metryka", "Umowa i prawa", "Materiały", "Wpływy i koszty", "Waterfall", "Rozliczenia"):
            self.assertContains(response, label)

        response = self.client.get(reverse("distribution:title_detail", args=[self.title.pk]))
        self.assertEqual(response.status_code, 200)
        for anchor in ("metadata", "rights", "materials", "finance", "waterfall", "settlements"):
            self.assertContains(response, f'id="{anchor}"')
        self.assertContains(response, "Rozlicz okres")

    def test_new_title_starts_workflow_outside_admin(self):
        response = self.client.post(reverse("distribution:title_create"), {
            "title_pl": "Nowy film workflow",
            "original_title": "Workflow Movie",
            "production_year": "2026",
            "status": "acquired",
            "producer": self.producer.pk,
            "acquisition_currency": Currency.PLN,
            "mg_advance": "0.00",
        })
        created = Title.objects.get(title_pl="Nowy film workflow")
        self.assertRedirects(response, reverse("distribution:title_detail", args=[created.pk]))

    def test_admin_save_stays_in_admin(self):
        response = self.client.post(reverse("admin:distribution_salesreport_add"), {
            "title": self.title.pk,
            "counterparty": self.cinema.pk,
            "sales_agreement": "",
            "exploitation_field": ExploitationField.CINEMA,
            "territory": "",
            "period_start": "2026-07-01",
            "period_end": "2026-07-31",
            "currency": Currency.PLN,
            "gross_revenue": "125.00",
            "deductions": "25.00",
            "vat_withholding": "0.00",
            "status": "checked",
            "source_reference": "ADMIN-RETURN-TEST",
            "_save": "Zapisz",
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("admin:distribution_salesreport_changelist"))

    def test_admin_uses_application_sidebar_and_settings_links(self):
        response = self.client.get(reverse("admin:index"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Panel administracyjny")
        self.assertNotContains(response, "FILMERP Panel Główny")
        self.assertContains(response, 'class="app-sidebar"')
        self.assertContains(response, "Panel administracyjny")
        self.assertContains(response, "Użytkownicy i role")
        self.assertContains(response, "Zmień hasło")
        self.assertContains(response, 'class="sidebar-logout-form"')
        self.assertContains(response, "Wyloguj")
        self.assertContains(response, "distribution/filmerp-sidebar.js")
        self.assertContains(response, "distribution/filmerp-favicon.svg")
        self.assertNotContains(response, ">DASHBOARD<")
        self.assertNotContains(response, 'id="user-tools"')
        self.assertContains(response, "distribution/filmerp-logo.svg")
        self.assertNotContains(response, "filter=id")

    def test_login_always_opens_dashboard_even_with_admin_next(self):
        self.client.logout()
        response = self.client.post(reverse("filmerp_login"), {
            "username": "finance",
            "password": "test-pass",
            "next": reverse("admin:index"),
        })
        self.assertRedirects(response, reverse("distribution:dashboard"))

    def test_admin_filters_are_collapsed_and_legacy_waterfall_is_hidden(self):
        response = self.client.get(reverse("admin:distribution_salesreport_changelist"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pokaż filtry")
        self.assertNotIn(WaterfallRecoupmentRule, admin.site._registry)
        self.assertNotIn(WaterfallRecoupmentItem, admin.site._registry)
        self.assertNotIn(WaterfallParticipant, admin.site._registry)
        index_response = self.client.get(reverse("admin:index"))
        self.assertContains(index_response, "Plany waterfall")
        self.assertContains(index_response, "Rozliczenia waterfall okresów")
        self.assertNotContains(index_response, "Kroki waterfall")
        self.assertNotContains(index_response, "Pozycje kalkulacji waterfall")

        cost_add_response = self.client.get(reverse("admin:distribution_cost_add"))
        self.assertEqual(cost_add_response.status_code, 200)
        self.assertContains(cost_add_response, "Podział procentowy")
        self.assertContains(cost_add_response, "allocation_cinema")

        cost_save_response = self.client.post(reverse("admin:distribution_cost_add"), {
            "title": self.title.pk,
            "category": CostCategory.PA,
            "supplier": "",
            "cost_date": "2026-06-15",
            "currency": Currency.PLN,
            "net_amount": "1000.00",
            "vat_amount": "230.00",
            "recoupable": "on",
            "scope_mode": CostScopeMode.ALLOCATED,
            "allocation_cinema": "60.00",
            "allocation_svod": "40.00",
            "invoice_file": "",
            "notes": "Podział z panelu admina",
            "_save": "Zapisz",
        })
        self.assertEqual(cost_save_response.status_code, 302)
        admin_cost = Cost.objects.get(notes="Podział z panelu admina")
        self.assertEqual(admin_cost.scope_fields, [ExploitationField.CINEMA, ExploitationField.SVOD])
        self.assertEqual(admin_cost.allocation_percentages, {
            ExploitationField.CINEMA: "60.00",
            ExploitationField.SVOD: "40.00",
        })

    def test_reports_show_period_settlements_from_current_waterfall(self):
        run = calculate_waterfall_run(WaterfallRun.objects.create(
            plan=self.plan,
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
        ))
        response = self.client.get(reverse("distribution:reports"), {
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
            "title": self.title.pk,
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rozliczenia waterfall okresów")
        self.assertContains(response, run.plan.name)
        self.assertNotContains(response, "Pozycje recoup.")
        export_response = self.client.get(reverse("distribution:reports_export_csv"), {
            "format": "xlsx",
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
            "title": self.title.pk,
        })
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_export_360_supports_xlsx_and_csv_zip(self):
        Cost.objects.create(
            title=self.title,
            category=CostCategory.PA,
            cost_date=date(2026, 6, 10),
            currency=Currency.PLN,
            net_amount=Decimal("200.00"),
            recoupable=True,
            scope_mode=CostScopeMode.SELECTED,
            scope_fields=[ExploitationField.CINEMA],
        )
        payload = {
            "export_all": "on",
            "export_format": "xlsx",
            "financial_scope": "period",
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
        }
        response = self.client.post(reverse("distribution:title_catalog_export"), payload)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        self.assertIn("Podsumowanie", workbook.sheetnames)
        self.assertIn("Koszty P&A", workbook.sheetnames)
        self.assertIn("Alokacje kosztów", workbook.sheetnames)
        self.assertEqual(workbook["Tytuły"]["B2"].value, self.title.title_pl)
        self.assertEqual(workbook["Koszty P&A"]["L2"].value, "Kino")

        payload["export_format"] = "csv_zip"
        response = self.client.post(reverse("distribution:title_catalog_export"), payload)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        with ZipFile(BytesIO(response.content)) as archive:
            self.assertIn("02_tytuly.csv", archive.namelist())
            self.assertIn("15_statementy.csv", archive.namelist())
            self.assertTrue(archive.read("02_tytuly.csv").startswith(b"\xef\xbb\xbf"))

    def test_contract_wizard_creates_versioned_agreement_and_waterfall(self):
        distributor = Counterparty.objects.create(name="FILMERP Dystrybucja")
        payload = {
            "title": self.title.pk,
            "contract_number": "ACQ/2026/7",
            "licensor": self.producer.pk,
            "distributor": distributor.pk,
            "signed_date": "2026-07-01",
            "rights_start": "2026-07-01",
            "rights_end": "2031-06-30",
            "currency": Currency.PLN,
            "mg_advance": "25000.00",
            "distributor_fee_percent": "15.00",
            "pa_recoupable": "on",
            "pa_cost_categories": [CostCategory.PA, CostCategory.PR],
            "applies_to_all_exploitation_fields": "on",
            "licensor_share_percent": "60.00",
            "status": "signed",
        }
        response = self.client.post(reverse("distribution:contract_waterfall_wizard"), payload)
        self.assertEqual(response.status_code, 302)
        agreement = AcquisitionAgreement.objects.get(contract_number="ACQ/2026/7")
        plan = WaterfallPlan.objects.get(name="Główny waterfall")
        self.assertEqual(agreement.mg_advance, Decimal("25000.00"))
        self.assertEqual(plan.status, "active")
        self.assertEqual(list(plan.steps.order_by("phase", "sort_order").values_list("name", flat=True)), [
            "Fee dystrybutora", "Zwrot kosztów P&A", "Zwrot MG",
            "Udział licencjodawcy", "Udział dystrybutora",
        ])

        response = self.client.post(reverse("distribution:contract_waterfall_wizard"), payload | {"contract_number": "ACQ/2026/8"})
        self.assertEqual(response.status_code, 302)
        plans = WaterfallPlan.objects.filter(name="Główny waterfall").order_by("version")
        self.assertEqual(list(plans.values_list("version", "status")), [(1, "archived"), (2, "active")])

    def test_workbench_simulates_finalizes_and_generates_pdf(self):
        response = self.client.get("/settlements/", self.workflow_payload())
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Miesieczny film")
        self.assertContains(response, "Przelicz okres")

        response = self.client.post("/settlements/", {**self.workflow_payload(), "action": "simulate"})
        self.assertEqual(response.status_code, 302)
        run = WaterfallRun.objects.get(plan=self.plan)
        self.assertEqual(run.allocated_amount, Decimal("500.00"))

        response = self.client.post("/settlements/", {
            **self.workflow_payload(),
            "action": "finalize_and_generate",
            "run": str(run.pk),
            "run_id": str(run.pk),
            "recipient_ids": [str(self.producer.pk)],
        })
        self.assertEqual(response.status_code, 302)
        run.refresh_from_db()
        statement = RoyaltyStatement.objects.get(waterfall_run=run, recipient=self.producer)
        self.assertEqual(run.status, WaterfallRunStatus.FINALIZED)
        self.assertEqual(statement.amount_due, Decimal("500.00"))
        self.assertTrue(statement.statement_file.name.endswith(".pdf"))
        with statement.statement_file.open("rb") as pdf_file:
            self.assertEqual(pdf_file.read(4), b"%PDF")

        original_snapshot = statement.calculation_snapshot.copy()
        SalesReport.objects.create(
            title=self.title,
            counterparty=self.cinema,
            exploitation_field=ExploitationField.CINEMA,
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            currency=Currency.PLN,
            gross_revenue=Decimal("9000.00"),
        )
        response = self.client.post("/settlements/", {
            **self.workflow_payload(),
            "action": "finalize_and_generate",
            "run_id": str(run.pk),
            "recipient_ids": [str(self.producer.pk)],
        })
        self.assertEqual(response.status_code, 302)
        statement.refresh_from_db()
        self.assertEqual(statement.calculation_snapshot, original_snapshot)
        self.assertEqual(statement.gross_revenue, Decimal("1000.00"))

    def test_invoice_upload_creates_cost_for_selected_title(self):
        invoice = SimpleUploadedFile("fv-001.pdf", b"%PDF-1.4 test invoice", content_type="application/pdf")
        response = self.client.post("/settlements/", {
            **self.workflow_payload(),
            "action": "upload_cost_invoice",
            "cost_date": "2026-06-15",
            "currency": Currency.PLN,
            "category": CostCategory.PA,
            "net_amount": "200.00",
            "vat_amount": "46.00",
            "recoupable": "on",
            "scope_mode": CostScopeMode.ALL,
            "invoice_file": invoice,
        })
        self.assertEqual(response.status_code, 302)
        cost = Cost.objects.get(title=self.title)
        self.assertEqual(cost.net_amount, Decimal("200.00"))
        self.assertTrue(cost.recoupable)
        self.assertTrue(cost.invoice_file.name.endswith(".pdf"))

    def test_document_center_blocks_duplicate_and_creates_cost(self):
        invoice_bytes = b"%PDF-1.4 test invoice document"
        response = self.client.post(reverse("distribution:document_center"), {
            "action": "upload",
            "source_file": SimpleUploadedFile("FV_2026_001.pdf", invoice_bytes, content_type="application/pdf"),
        })
        self.assertEqual(response.status_code, 302)
        document = DocumentInboxItem.objects.get()
        self.assertEqual(document.document_type, DocumentType.COST_INVOICE)
        self.assertEqual(document.status, DocumentStatus.NEEDS_REVIEW)

        response = self.client.post(reverse("distribution:document_center"), {
            "action": "upload",
            "source_file": SimpleUploadedFile("kopia.pdf", invoice_bytes, content_type="application/pdf"),
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(DocumentInboxItem.objects.count(), 1)

        response = self.client.post(reverse("distribution:document_center"), {
            "action": "create_cost",
            "document_id": str(document.pk),
            "title": str(self.title.pk),
            "supplier_name": "Studio reklamowe",
            "cost_date": "2026-06-12",
            "category": CostCategory.PA,
            "currency": Currency.PLN,
            "net_amount": "500.00",
            "vat_amount": "115.00",
            "recoupable": "on",
            "scope_mode": CostScopeMode.ALL,
        })
        self.assertEqual(response.status_code, 302)
        document.refresh_from_db()
        self.assertEqual(document.status, DocumentStatus.PROCESSED)
        self.assertEqual(document.cost.net_amount, Decimal("500.00"))
        self.assertEqual(document.cost.title, self.title)
        self.assertEqual(Cost.objects.filter(title=self.title).count(), 1)

    def test_document_center_imports_selected_cinema_report_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = self.title.title_pl
        sheet["E1"] = "01.06.2026 - 30.06.2026"
        sheet["E2"] = "Widzow"
        sheet["F2"] = "Brutto"
        sheet["G2"] = "Netto"
        sheet["C3"] = "Warszawa"
        sheet["D3"] = self.cinema.name
        sheet["E3"] = 100
        sheet["F3"] = 1000
        buffer = BytesIO()
        workbook.save(buffer)

        response = self.client.post(reverse("distribution:document_center"), {
            "action": "upload",
            "source_file": SimpleUploadedFile(
                "raport_widzowie.xlsx",
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        })
        self.assertEqual(response.status_code, 302)
        document = DocumentInboxItem.objects.get()
        self.assertEqual(document.document_type, DocumentType.CINEMA_REPORT)
        row = document.cinema_import.rows.get()
        self.assertEqual(row.title, self.title)
        self.assertEqual(row.admissions, 100)

        response = self.client.post(reverse("distribution:document_center"), {
            "action": "approve_report_rows",
            "document_id": str(document.pk),
            "row_ids": [str(row.pk)],
        })
        self.assertEqual(response.status_code, 302)
        document.refresh_from_db()
        row.refresh_from_db()
        self.assertEqual(document.status, DocumentStatus.PROCESSED)
        self.assertIsNotNone(row.booking_id)
        self.assertTrue(SalesReport.objects.filter(source_reference=f"cinema-import-row-{row.pk}").exists())
