import csv
from io import BytesIO
from datetime import timedelta
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.views import LoginView
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .avails import check_availability
from .cinema_imports import approve_import_rows, parse_cinema_report_import
from .documents import analyze_document, ingest_document
from .forms import (
    CinemaReportUploadForm,
    CostInvoiceUploadForm,
    DocumentClassificationForm,
    DocumentCostForm,
    DocumentUploadForm,
    ContractWaterfallWizardForm,
    TitleSetupForm,
    TitleCatalogExportForm,
)
from .catalog_export import build_catalog_export, export_catalog_csv_zip, export_catalog_xlsx
from .contract_wizard import create_contract_waterfall
from .pdf import build_royalty_statement_pdf
from .models import (
    AcquisitionAgreement,
    CinemaBooking,
    CinemaReportImport,
    Cost,
    Counterparty,
    Currency,
    DeliveryStatus,
    DocumentInboxItem,
    DocumentStatus,
    DocumentType,
    ExploitationField,
    LanguageVersion,
    RightsIssue,
    RightsStatus,
    RightsWindow,
    RoyaltyStatement,
    ReportStatus,
    SalesAgreement,
    SalesReport,
    StatementStatus,
    Territory,
    Title,
    TitleMaterial,
    WaterfallPlan,
    WaterfallPlanStatus,
    WaterfallRun,
    WaterfallRunStatus,
)
from .settlements import create_statement_documents
from .waterfall_engine import calculate_waterfall_run, finalize_waterfall_run


class DashboardLoginView(LoginView):
    template_name = "admin/login.html"
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse("distribution:dashboard")


@login_required
def dashboard(request):
    today = timezone.localdate()
    expiring_until = today + timedelta(days=90)
    open_statements = RoyaltyStatement.objects.exclude(status=StatementStatus.PAID)
    query = request.GET.get("q", "").strip()

    titles = Title.objects.select_related("producer").annotate(
        agreements_count=Count("acquisition_agreements", distinct=True),
        rights_count=Count("rights_windows", distinct=True),
        materials_count=Count("materials", distinct=True),
        reports_count=Count("sales_reports", distinct=True),
        costs_count=Count("costs", distinct=True),
        statements_count=Count("royalty_statements", distinct=True),
        active_waterfall_steps=Count(
            "waterfall_plans__steps",
            filter=Q(waterfall_plans__status=WaterfallPlanStatus.ACTIVE, waterfall_plans__steps__active=True),
            distinct=True,
        ),
        finalized_runs_count=Count(
            "waterfall_plans__runs",
            filter=Q(waterfall_plans__runs__status=WaterfallRunStatus.FINALIZED),
            distinct=True,
        ),
    ).order_by("title_pl")
    if query:
        titles = titles.filter(Q(title_pl__icontains=query) | Q(original_title__icontains=query))

    title_rows = []
    for title in titles[:100]:
        stages = [
            {"key": "metadata", "label": "Metryka", "ready": bool(title.production_year and title.producer_id)},
            {"key": "rights", "label": "Umowa i prawa", "ready": bool(title.agreements_count and title.rights_count)},
            {"key": "materials", "label": "Materiały", "ready": bool(title.materials_count)},
            {"key": "finance", "label": "Wpływy i koszty", "ready": bool(title.reports_count or title.costs_count)},
            {"key": "waterfall", "label": "Waterfall", "ready": bool(title.active_waterfall_steps)},
            {"key": "settlements", "label": "Rozliczenia", "ready": bool(title.finalized_runs_count)},
        ]
        done = sum(stage["ready"] for stage in stages)
        next_stage = next((stage for stage in stages if not stage["ready"]), stages[-1])
        title_rows.append({
            "title": title,
            "stages": stages,
            "done": done,
            "percent": round(done / len(stages) * 100),
            "next_label": next_stage["label"],
            "next_anchor": next_stage["key"],
        })

    revenue_by_title = (
        SalesReport.objects.values("title__title_pl", "currency")
        .annotate(gross=Sum("gross_revenue"), count=Count("id"))
        .order_by("-gross")[:10]
    )
    revenue_by_field = (
        SalesReport.objects.values("exploitation_field", "currency")
        .annotate(gross=Sum("gross_revenue"), count=Count("id"))
        .order_by("-gross")[:10]
    )

    currencies = sorted(set(SalesReport.objects.values_list("currency", flat=True)) | set(Cost.objects.values_list("currency", flat=True)) | set(open_statements.values_list("currency", flat=True)))
    financial_totals = []
    for currency in currencies:
        gross = SalesReport.objects.filter(currency=currency).aggregate(total=Sum("gross_revenue"))["total"] or 0
        costs = Cost.objects.filter(currency=currency).aggregate(total=Sum("net_amount"))["total"] or 0
        statements = [statement for statement in open_statements if statement.currency == currency]
        financial_totals.append({
            "currency": currency,
            "gross_revenue": gross,
            "costs": costs,
            "open_statements": sum((statement.amount_due for statement in statements), 0),
        })

    context = {
        "query": query,
        "title_rows": title_rows,
        "titles_count": Title.objects.count(),
        "active_rights_count": RightsWindow.objects.exclude(status__in=[RightsStatus.EXPIRED, RightsStatus.CANCELLED]).count(),
        "expiring_rights": RightsWindow.objects.filter(date_to__gte=today, date_to__lte=expiring_until).order_by("date_to")[:10],
        "open_issues_count": RightsIssue.objects.filter(resolved=False).count(),
        "open_issues": RightsIssue.objects.filter(resolved=False).select_related("rights_window", "rights_window__title")[:10],
        "unpaid_agreements": SalesAgreement.objects.filter(invoice_paid=False, payment_due_date__lt=today).select_related("title", "licensee")[:10],
        "financial_totals": financial_totals,
        "open_statements_count": open_statements.count(),
        "revenue_by_title": revenue_by_title,
        "revenue_by_field": revenue_by_field,
    }
    return render(request, "distribution/dashboard.html", context)


@login_required
def title_list(request):
    query = request.GET.get("q", "").strip()
    titles_qs = Title.objects.select_related("producer").order_by("title_pl")
    if query:
        titles_qs = titles_qs.filter(
            Q(title_pl__icontains=query)
            | Q(original_title__icontains=query)
            | Q(ean__icontains=query)
            | Q(director__icontains=query)
            | Q(cast__icontains=query)
        )
    titles_qs = titles_qs[:200]
    context = {
        "query": query,
        "titles": titles_qs,
    }
    return render(request, "distribution/title_list.html", context)


@login_required
def title_setup(request, pk=None):
    title = get_object_or_404(Title, pk=pk) if pk else None
    permission = "distribution.change_title" if title else "distribution.add_title"
    if not request.user.has_perm(permission):
        raise PermissionDenied
    form = TitleSetupForm(request.POST or None, instance=title)
    if request.method == "POST" and form.is_valid():
        title = form.save()
        messages.success(request, "Metryka tytułu została zapisana. Przejdź do kolejnego etapu.")
        return redirect("distribution:title_detail", pk=title.pk)
    return render(request, "distribution/title_setup.html", {"form": form, "title": title})


@login_required
def title_detail(request, pk):
    title = get_object_or_404(Title.objects.select_related("producer"), pk=pk)
    today = timezone.localdate()

    rights_windows = (
        title.rights_windows.select_related("counterparty", "acquisition_agreement", "sales_agreement")
        .prefetch_related("territories", "language_versions")
        .order_by("exploitation_field", "date_from")
    )
    sales_reports = title.sales_reports.select_related("counterparty", "territory").order_by("-period_end")[:20]
    costs = title.costs.select_related("supplier").order_by("-cost_date")[:20]
    cinema_bookings = title.cinema_bookings.select_related("cinema").order_by("-date_from")[:20]
    royalty_statements = title.royalty_statements.select_related("recipient").order_by("-period_end")[:10]
    materials = title.materials.select_related("language_version", "supplier").order_by("asset_type", "due_date")
    issues = RightsIssue.objects.filter(rights_window__title=title, resolved=False).select_related("rights_window", "conflicting_window")
    agreements = title.acquisition_agreements.select_related("licensor").prefetch_related("territories").order_by("-signed_date", "-created_at")
    waterfall_plans = title.waterfall_plans.prefetch_related("steps__beneficiary").order_by("-version")
    active_plan = waterfall_plans.filter(status=WaterfallPlanStatus.ACTIVE).first() or waterfall_plans.first()
    active_steps = active_plan.steps.filter(active=True).select_related("beneficiary").order_by("phase", "sort_order") if active_plan else []

    gross_box_office = title.cinema_bookings.aggregate(total=Sum("box_office_gross"))["total"] or 0
    admissions = title.cinema_bookings.aggregate(total=Sum("admissions"))["total"] or 0
    screenings = title.cinema_bookings.aggregate(total=Sum("screenings"))["total"] or 0
    required_materials = materials.filter(required_for_release=True)
    open_materials = required_materials.exclude(status__in=[DeliveryStatus.READY, DeliveryStatus.SENT, DeliveryStatus.ACCEPTED])
    overdue_materials = [material for material in open_materials if material.is_overdue]
    title_documents = title.inbox_documents.order_by("-created_at")[:10]
    finalized_runs_count = WaterfallRun.objects.filter(plan__title=title, status=WaterfallRunStatus.FINALIZED).count()
    workflow_stages = [
        {"key": "metadata", "label": "Metryka", "ready": bool(title.production_year and title.producer_id), "detail": "rok i producent"},
        {"key": "rights", "label": "Umowa i prawa", "ready": agreements.exists() and rights_windows.exists(), "detail": f"{agreements.count()} umów, {rights_windows.count()} praw"},
        {"key": "materials", "label": "Materiały", "ready": materials.exists(), "detail": f"{materials.count()} pozycji"},
        {"key": "finance", "label": "Wpływy i koszty", "ready": sales_reports.exists() or costs.exists(), "detail": f"{sales_reports.count()} raportów, {costs.count()} kosztów"},
        {"key": "waterfall", "label": "Waterfall", "ready": bool(active_plan and active_steps), "detail": f"{len(active_steps)} kroków" if active_plan else "brak planu"},
        {"key": "settlements", "label": "Rozliczenia", "ready": bool(finalized_runs_count), "detail": f"{finalized_runs_count} zamkniętych okresów"},
    ]

    context = {
        "title": title,
        "today": today,
        "rights_windows": rights_windows,
        "sales_reports": sales_reports,
        "costs": costs,
        "cinema_bookings": cinema_bookings,
        "royalty_statements": royalty_statements,
        "materials": materials,
        "issues": issues,
        "agreements": agreements,
        "active_plan": active_plan,
        "active_steps": active_steps,
        "title_documents": title_documents,
        "workflow_stages": workflow_stages,
        "workflow_done": sum(stage["ready"] for stage in workflow_stages),
        "gross_box_office": gross_box_office,
        "admissions": admissions,
        "screenings": screenings,
        "required_materials_count": required_materials.count(),
        "open_materials_count": open_materials.count(),
        "overdue_materials_count": len(overdue_materials),
    }
    return render(request, "distribution/title_detail.html", context)


@login_required
@permission_required(("distribution.add_acquisitionagreement", "distribution.add_waterfallplan"), raise_exception=True)
def contract_waterfall_wizard(request):
    title = None
    if request.GET.get("title"):
        title = get_object_or_404(Title, pk=request.GET["title"])
    if request.method == "POST":
        form = ContractWaterfallWizardForm(request.POST)
        if form.is_valid():
            agreement, plan = create_contract_waterfall(form.cleaned_data)
            messages.success(request, f"Utworzono umowę i aktywny waterfall v{plan.version} z {plan.steps.count()} krokami.")
            return redirect(f"{reverse('distribution:settlement_workbench')}?title={agreement.title_id}&currency={plan.currency}&plan={plan.pk}")
    else:
        form = ContractWaterfallWizardForm(title=title)
    return render(request, "distribution/contract_waterfall_wizard.html", {
        "form": form,
        "selected_title": title,
    })


@login_required
def avails(request):
    result = None
    errors = []
    selected_title = None
    selected_territories = Territory.objects.none()
    selected_languages = LanguageVersion.objects.none()

    title_id = request.GET.get("title") or ""
    exploitation_field = request.GET.get("exploitation_field") or ""
    territory_ids = [item for item in request.GET.getlist("territories") if item]
    language_ids = [item for item in request.GET.getlist("languages") if item]
    date_from_raw = request.GET.get("date_from") or ""
    date_to_raw = request.GET.get("date_to") or ""

    if any([title_id, exploitation_field, territory_ids, language_ids, date_from_raw, date_to_raw]):
        date_from = parse_date(date_from_raw)
        date_to = parse_date(date_to_raw)
        if not title_id:
            errors.append("Wybierz tytuł.")
        if not exploitation_field:
            errors.append("Wybierz pole eksploatacji.")
        if not date_from or not date_to:
            errors.append("Podaj poprawny zakres dat.")
        elif date_from > date_to:
            errors.append("Data od nie może być późniejsza niż data do.")
        if title_id:
            selected_title = get_object_or_404(Title, pk=title_id)
        selected_territories = Territory.objects.filter(pk__in=territory_ids).order_by("name")
        selected_languages = LanguageVersion.objects.filter(pk__in=language_ids).order_by("name")

        if not errors:
            result = check_availability(
                title=selected_title,
                exploitation_field=exploitation_field,
                territories=selected_territories,
                languages=selected_languages,
                date_from=date_from,
                date_to=date_to,
            )

    context = {
        "result": result,
        "errors": errors,
        "titles": Title.objects.order_by("title_pl"),
        "territories": Territory.objects.order_by("name"),
        "languages": LanguageVersion.objects.order_by("name"),
        "exploitation_fields": ExploitationField.choices,
        "filters": {
            "title": title_id,
            "exploitation_field": exploitation_field,
            "territories": territory_ids,
            "languages": language_ids,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
        },
        "selected_title": selected_title,
        "selected_territories": selected_territories,
        "selected_languages": selected_languages,
    }
    return render(request, "distribution/avails.html", context)


def _previous_month_period():
    today = timezone.localdate()
    last_day_previous_month = today.replace(day=1) - timedelta(days=1)
    return last_day_previous_month.replace(day=1), last_day_previous_month


def _document_center_url(document_id=""):
    url = reverse("distribution:document_center")
    return f"{url}?document={document_id}" if document_id else url


@login_required
@permission_required("distribution.view_documentinboxitem", raise_exception=True)
def document_center(request):
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "upload":
            if not request.user.has_perm("distribution.add_documentinboxitem"):
                raise PermissionDenied
            form = DocumentUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    document, created = ingest_document(request.FILES["source_file"], request.user)
                    if created:
                        messages.success(request, "Dokument wczytano i przekazano do weryfikacji.")
                    else:
                        messages.warning(request, "Ten sam plik jest już w Centrum dokumentów. Otwieram istniejący rekord.")
                    return redirect(_document_center_url(document.pk))
                except ValueError as exc:
                    messages.error(request, str(exc))
            else:
                messages.error(request, "Wybierz plik PDF, XLSX albo obraz faktury.")
            return redirect(_document_center_url())

        document = get_object_or_404(
            DocumentInboxItem.objects.select_related("cinema_import", "cost"),
            pk=request.POST.get("document_id"),
        )
        if not request.user.has_perm("distribution.change_documentinboxitem"):
            raise PermissionDenied

        if action == "classify":
            if document.status == DocumentStatus.PROCESSED:
                messages.error(request, "Zaksięgowanego dokumentu nie można ponownie sklasyfikować.")
            else:
                form = DocumentClassificationForm(request.POST, instance=document)
                if form.is_valid():
                    chosen_type = form.cleaned_data["document_type"]
                    try:
                        analyze_document(document, forced_type=chosen_type)
                        messages.success(request, "Zmieniono rodzaj dokumentu i ponowiono analizę.")
                    except ValueError as exc:
                        messages.error(request, str(exc))
            return redirect(_document_center_url(document.pk))

        if action == "apply_report_title":
            if not document.cinema_import_id:
                messages.error(request, "Ten dokument nie ma rozpoznanych wierszy raportu.")
            else:
                row_ids = request.POST.getlist("row_ids")
                if not row_ids:
                    messages.error(request, "Zaznacz co najmniej jeden wiersz.")
                else:
                    title = get_object_or_404(Title, pk=request.POST.get("title_id"))
                    updated = document.cinema_import.rows.filter(pk__in=row_ids).exclude(status="imported").update(title=title)
                    document.title = title
                    document.save(update_fields=["title", "updated_at"])
                    messages.success(request, f"Przypisano tytuł do {updated} wierszy.")
            return redirect(_document_center_url(document.pk))

        if action == "approve_report_rows":
            if not request.user.has_perm("distribution.add_cinemabooking") or not request.user.has_perm("distribution.add_salesreport"):
                raise PermissionDenied
            if not document.cinema_import_id:
                messages.error(request, "Ten dokument nie ma rozpoznanych wierszy raportu.")
            else:
                row_ids = request.POST.getlist("row_ids")
                rows = document.cinema_import.rows.filter(pk__in=row_ids)
                if not row_ids:
                    messages.error(request, "Zaznacz co najmniej jeden wiersz.")
                else:
                    imported, skipped = approve_import_rows(rows)
                    document.cinema_import.refresh_from_db()
                    if document.cinema_import.status == "imported":
                        document.status = DocumentStatus.PROCESSED
                        document.processed_at = timezone.now()
                    document.reviewed_by = request.user
                    document.save(update_fields=["status", "processed_at", "reviewed_by", "updated_at"])
                    if skipped:
                        messages.warning(request, f"Utworzono {imported} bookingów. Pominięto {skipped} wierszy wymagających poprawy.")
                    else:
                        messages.success(request, f"Utworzono bookingi i raporty sprzedaży: {imported}.")
            return redirect(_document_center_url(document.pk))

        if action == "create_cost":
            if not request.user.has_perm("distribution.add_cost"):
                raise PermissionDenied
            if document.cost_id:
                messages.warning(request, "Ta faktura została już zaksięgowana jako koszt.")
            else:
                form = DocumentCostForm(request.POST, document=document)
                if form.is_valid():
                    cost = form.save()
                    document.cost = cost
                    document.title = cost.title
                    document.counterparty = cost.supplier
                    document.status = DocumentStatus.PROCESSED
                    document.reviewed_by = request.user
                    document.processed_at = timezone.now()
                    document.save(update_fields=["cost", "title", "counterparty", "status", "reviewed_by", "processed_at", "updated_at"])
                    messages.success(request, f"Faktura została zaksięgowana jako koszt {cost.net_amount} {cost.currency}.")
                else:
                    errors = " ".join(error for field_errors in form.errors.values() for error in field_errors)
                    messages.error(request, f"Nie utworzono kosztu. {errors}")
            return redirect(_document_center_url(document.pk))

        if action == "reject":
            has_imported_rows = document.cinema_import_id and document.cinema_import.rows.filter(status="imported").exists()
            if document.cost_id or has_imported_rows:
                messages.error(request, "Dokument ma już zaksięgowane dane i nie może zostać odrzucony.")
            else:
                document.status = DocumentStatus.REJECTED
                document.reviewed_by = request.user
                document.save(update_fields=["status", "reviewed_by", "updated_at"])
                messages.success(request, "Dokument oznaczono jako odrzucony.")
            return redirect(_document_center_url(document.pk))

    base_documents = DocumentInboxItem.objects.select_related(
        "title", "counterparty", "uploaded_by", "reviewed_by", "cinema_import", "cost"
    )
    status_filter = request.GET.get("status", "")
    queue = base_documents
    if status_filter in DocumentStatus.values:
        queue = queue.filter(status=status_filter)

    selected_document = None
    selected_id = request.GET.get("document")
    if selected_id:
        selected_document = get_object_or_404(base_documents, pk=selected_id)
    elif queue.exists():
        selected_document = queue.first()

    report_rows = selected_document.cinema_import.rows.select_related("title", "cinema", "booking") if selected_document and selected_document.cinema_import_id else None
    classification_form = DocumentClassificationForm(instance=selected_document) if selected_document else None
    cost_form = None
    if selected_document and selected_document.document_type == DocumentType.COST_INVOICE and not selected_document.cost_id:
        cost_form = DocumentCostForm(document=selected_document)

    context = {
        "documents": queue,
        "selected_document": selected_document,
        "report_rows": report_rows,
        "upload_form": DocumentUploadForm(),
        "classification_form": classification_form,
        "cost_form": cost_form,
        "titles": Title.objects.order_by("title_pl"),
        "status_filter": status_filter,
        "document_statuses": DocumentStatus.choices,
        "document_count": base_documents.count(),
        "review_count": base_documents.filter(status=DocumentStatus.NEEDS_REVIEW).count(),
        "processed_count": base_documents.filter(status=DocumentStatus.PROCESSED).count(),
    }
    return render(request, "distribution/document_center.html", context)


def _settlement_url(*, title_id="", plan_id="", period_start="", period_end="", currency="", run_id="", import_id=""):
    params = {
        "title": title_id,
        "plan": plan_id,
        "date_from": period_start,
        "date_to": period_end,
        "currency": currency,
        "run": run_id,
        "import": import_id,
    }
    query = urlencode({key: value for key, value in params.items() if value})
    return f"/settlements/?{query}" if query else "/settlements/"


@login_required
def settlement_workbench(request):
    default_start, default_end = _previous_month_period()
    values = request.POST if request.method == "POST" else request.GET
    run_id = values.get("run") or ""
    run = None
    if run_id:
        run = get_object_or_404(
            WaterfallRun.objects.select_related("plan", "plan__title", "finalized_by").prefetch_related(
                "lines__cost_allocations__cost",
            ),
            pk=run_id,
        )

    selected_title = run.plan.title if run else None
    if not selected_title and values.get("title"):
        selected_title = get_object_or_404(Title, pk=values.get("title"))

    period_start = run.period_start if run else (parse_date(values.get("date_from") or "") or default_start)
    period_end = run.period_end if run else (parse_date(values.get("date_to") or "") or default_end)
    if period_start > period_end:
        period_start, period_end = period_end, period_start
    currency = run.plan.currency if run else (values.get("currency") or Currency.PLN)

    plans = WaterfallPlan.objects.none()
    selected_plan = run.plan if run else None
    if selected_title:
        plans = selected_title.waterfall_plans.filter(currency=currency).prefetch_related("steps").order_by("-version")
        if not selected_plan and values.get("plan"):
            selected_plan = get_object_or_404(plans, pk=values.get("plan"))
        if not selected_plan:
            selected_plan = plans.filter(status=WaterfallPlanStatus.ACTIVE).first() or plans.first()

    redirect_kwargs = {
        "title_id": selected_title.pk if selected_title else "",
        "plan_id": selected_plan.pk if selected_plan else "",
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "currency": currency,
        "run_id": run.pk if run else "",
        "import_id": values.get("import") or "",
    }

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "upload_cinema_report":
            form = CinemaReportUploadForm(request.POST, request.FILES)
            if form.is_valid():
                report_import = form.save(commit=False)
                report_import.original_filename = request.FILES["source_file"].name
                report_import.save()
                try:
                    parsed = parse_cinema_report_import(report_import)
                    messages.success(request, f"Rozpoznano {parsed} wierszy. Sprawdz je przed akceptacja.")
                except ValueError as exc:
                    messages.error(request, str(exc))
                redirect_kwargs["import_id"] = report_import.pk
            else:
                messages.error(request, "Wybierz raport kina w formacie PDF albo XLSX.")
            return redirect(_settlement_url(**redirect_kwargs))

        if action == "approve_cinema_import":
            report_import = get_object_or_404(CinemaReportImport, pk=request.POST.get("import_id"))
            imported, skipped = approve_import_rows(report_import.rows.all())
            if skipped:
                messages.warning(request, f"Utworzono {imported} bookingow. {skipped} wierszy nadal wymaga uzupelnienia.")
            else:
                messages.success(request, f"Utworzono bookingi i raporty sprzedazy: {imported}.")
            redirect_kwargs["import_id"] = report_import.pk
            return redirect(_settlement_url(**redirect_kwargs))

        if action == "upload_cost_invoice":
            if not selected_title:
                messages.error(request, "Najpierw wybierz tytul.")
            else:
                form = CostInvoiceUploadForm(request.POST, request.FILES, title=selected_title, currency=currency)
                if form.is_valid():
                    cost = form.save()
                    messages.success(request, f"Faktura zostala przypisana jako koszt {cost.net_amount} {cost.currency}.")
                else:
                    errors = " ".join(error for field_errors in form.errors.values() for error in field_errors)
                    messages.error(request, f"Nie zapisano faktury. {errors}")
            return redirect(_settlement_url(**redirect_kwargs))

        if action == "simulate":
            if not selected_plan:
                messages.error(request, "Najpierw utworz i wybierz plan waterfall dla tytulu.")
            elif not selected_plan.steps.filter(active=True).exists():
                messages.error(request, "Plan waterfall nie ma aktywnych krokow.")
            else:
                run = WaterfallRun.objects.filter(
                    plan=selected_plan,
                    period_start=period_start,
                    period_end=period_end,
                    status=WaterfallRunStatus.DRAFT,
                ).order_by("-created_at").first()
                if run is None:
                    run = WaterfallRun.objects.create(plan=selected_plan, period_start=period_start, period_end=period_end)
                calculate_waterfall_run(run)
                redirect_kwargs["run_id"] = run.pk
                messages.success(request, "Okres został przeliczony. Wynik jest roboczy i nie został jeszcze zatwierdzony.")
            return redirect(_settlement_url(**redirect_kwargs))

        if action == "finalize_and_generate":
            run = get_object_or_404(WaterfallRun.objects.select_related("plan"), pk=request.POST.get("run_id"))
            recipient_ids = request.POST.getlist("recipient_ids")
            allowed_ids = set(run.plan.steps.filter(active=True, beneficiary__isnull=False).values_list("beneficiary_id", flat=True))
            selected_recipient_ids = {
                int(value) for value in recipient_ids if str(value).isdigit()
            }
            if not (selected_recipient_ids & allowed_ids):
                messages.error(request, "Wybierz co najmniej jednego odbiorce statementu.")
            else:
                try:
                    if run.status == WaterfallRunStatus.DRAFT:
                        calculate_waterfall_run(run)
                        finalize_waterfall_run(run, request.user)
                    elif run.status != WaterfallRunStatus.FINALIZED:
                        raise ValidationError("Anulowanego rozliczenia nie mozna wygenerowac.")
                    statements = create_statement_documents(run, recipient_ids)
                    messages.success(request, f"Rozliczenie zatwierdzone. Utworzono statementy PDF: {len(statements)}.")
                except ValidationError as exc:
                    messages.error(request, " ".join(exc.messages))
            redirect_kwargs["run_id"] = run.pk
            return redirect(_settlement_url(**redirect_kwargs))

    agreements = AcquisitionAgreement.objects.none()
    sales_reports = SalesReport.objects.none()
    costs = Cost.objects.none()
    if selected_title:
        agreements = selected_title.acquisition_agreements.select_related("licensor").order_by("-signed_date", "-created_at")
        sales_reports = selected_title.sales_reports.filter(
            currency=currency,
            period_start__lte=period_end,
            period_end__gte=period_start,
        ).exclude(status=ReportStatus.REJECTED).select_related("counterparty", "territory").order_by("period_start", "counterparty__name")
        costs = selected_title.costs.filter(
            currency=currency,
            cost_date__gte=period_start,
            cost_date__lte=period_end,
        ).select_related("supplier").order_by("cost_date", "supplier__name")

    agreement = agreements.first()
    active_steps = selected_plan.steps.filter(active=True).select_related("beneficiary") if selected_plan else []
    steps_count = active_steps.count() if selected_plan else 0
    recipients = []
    if selected_plan:
        recipient_ids = list(active_steps.filter(beneficiary__isnull=False).values_list("beneficiary_id", flat=True).distinct())
        for recipient in Counterparty.objects.filter(pk__in=recipient_ids).order_by("name"):
            allocated = run.lines.filter(beneficiary=recipient).aggregate(total=Sum("allocated_amount"))["total"] if run else 0
            recipients.append({"counterparty": recipient, "allocated": allocated or 0})

    readiness = [
        {"label": "Umowa nabycia", "ready": bool(agreement), "warning": not agreement, "detail": str(agreement) if agreement else "Brak umowy przypisanej do tytulu"},
        {"label": "Warunki waterfall", "ready": bool(selected_plan and steps_count), "warning": False, "detail": f"{selected_plan.name}, v{selected_plan.version}, {steps_count} krokow" if selected_plan and steps_count else "Brak kompletnego planu"},
        {"label": "Raporty za okres", "ready": sales_reports.exists(), "warning": False, "detail": f"{sales_reports.count()} raportow" if selected_title else "Najpierw wybierz tytul"},
        {"label": "Koszty i faktury", "ready": costs.exists(), "warning": not costs.exists(), "detail": f"{costs.count()} kosztow, {costs.exclude(invoice_file='').count()} z plikiem FV" if selected_title else "Najpierw wybierz tytul"},
        {"label": "Odbiorcy statementow", "ready": bool(recipients), "warning": False, "detail": f"{len(recipients)} odbiorcow" if recipients else "Brak beneficjentow w planie"},
    ]

    selected_import = None
    import_id = request.GET.get("import") or ""
    if import_id:
        selected_import = get_object_or_404(CinemaReportImport.objects.prefetch_related("rows__title", "rows__cinema"), pk=import_id)
    recent_imports = CinemaReportImport.objects.prefetch_related("rows").order_by("-created_at")[:6]
    statements = run.royalty_statements.select_related("recipient").order_by("recipient__name") if run else RoyaltyStatement.objects.none()
    invoice_form = CostInvoiceUploadForm(title=selected_title, currency=currency) if selected_title else None

    context = {
        "titles": Title.objects.order_by("title_pl"),
        "currencies": Currency.choices,
        "selected_title": selected_title,
        "selected_plan": selected_plan,
        "plans": plans,
        "period_start": period_start,
        "period_end": period_end,
        "currency": currency,
        "agreement": agreement,
        "agreements": agreements,
        "active_steps": active_steps,
        "sales_reports": sales_reports,
        "costs": costs,
        "readiness": readiness,
        "run": run,
        "recipients": recipients,
        "statements": statements,
        "cinema_upload_form": CinemaReportUploadForm(),
        "invoice_form": invoice_form,
        "selected_import": selected_import,
        "recent_imports": recent_imports,
        "can_simulate": bool(selected_plan and steps_count),
    }
    return render(request, "distribution/settlement_workbench.html", context)


def _filtered_statements(request):
    statements = RoyaltyStatement.objects.select_related("title", "recipient").order_by("-period_end", "title__title_pl")
    filters = {
        "status": request.GET.get("status") or "",
        "title_id": request.GET.get("title") or "",
        "recipient_id": request.GET.get("recipient") or "",
        "date_from": request.GET.get("date_from") or "",
        "date_to": request.GET.get("date_to") or "",
    }
    if filters["status"]:
        statements = statements.filter(status=filters["status"])
    if filters["title_id"]:
        statements = statements.filter(title_id=filters["title_id"])
    if filters["recipient_id"]:
        statements = statements.filter(recipient_id=filters["recipient_id"])
    if filters["date_from"]:
        statements = statements.filter(period_end__gte=filters["date_from"])
    if filters["date_to"]:
        statements = statements.filter(period_start__lte=filters["date_to"])
    return statements, filters


@login_required
def statement_center(request):
    if request.method == "POST":
        action = request.POST.get("action")
        statement_ids = request.POST.getlist("statement_ids")
        statements = RoyaltyStatement.objects.filter(pk__in=statement_ids).select_related("title", "recipient")
        if not statement_ids:
            messages.warning(request, "Zaznacz przynajmniej jeden statement.")
            return redirect("distribution:statement_center")

        if action == "generate_pdf":
            generated = 0
            for statement in statements:
                statement.freeze_calculation(lock=True)
                pdf_file = build_royalty_statement_pdf(statement)
                statement.statement_file.save(pdf_file.name, pdf_file, save=True)
                generated += 1
            messages.success(request, f"Wygenerowano PDF: {generated}.")
        elif action == "mark_sent":
            for statement in statements:
                statement.freeze_calculation(lock=True)
            updated = statements.update(status=StatementStatus.SENT, sent_at=timezone.localdate())
            messages.success(request, f"Oznaczono jako wyslane: {updated}.")
        elif action == "mark_approved":
            for statement in statements:
                statement.freeze_calculation(lock=True)
            updated = statements.update(status=StatementStatus.APPROVED)
            messages.success(request, f"Oznaczono jako zaakceptowane: {updated}.")
        elif action == "mark_paid":
            for statement in statements:
                statement.freeze_calculation(lock=True)
            updated = statements.update(status=StatementStatus.PAID, paid_at=timezone.localdate())
            messages.success(request, f"Oznaczono jako oplacone: {updated}.")
        elif action == "mark_disputed":
            updated = statements.update(status=StatementStatus.DISPUTED)
            messages.success(request, f"Oznaczono jako sporne: {updated}.")
        else:
            messages.warning(request, "Nieznana akcja.")
        return redirect("distribution:statement_center")

    statements, filters = _filtered_statements(request)
    if request.GET.get("export") == "xlsx":
        return _statement_center_export_xlsx(statements, filters)

    statement_list = list(statements[:300])
    amount_totals = []
    for currency in sorted({statement.currency for statement in statement_list}):
        amount_totals.append({
            "currency": currency,
            "amount_due": sum((statement.amount_due for statement in statement_list if statement.currency == currency), 0),
            "open_amount": sum((statement.amount_due for statement in statement_list if statement.currency == currency and statement.status != StatementStatus.PAID), 0),
        })
    status_counts = {
        row["status"]: row["count"]
        for row in statements.values("status").annotate(count=Count("id"))
    }
    status_count_rows = [
        {"value": value, "label": label, "count": status_counts.get(value, 0)}
        for value, label in StatementStatus.choices
    ]

    context = {
        "statements": statement_list,
        "filters": filters,
        "titles": Title.objects.order_by("title_pl"),
        "recipients": Counterparty.objects.order_by("name"),
        "statement_statuses": StatementStatus.choices,
        "status_counts": status_counts,
        "status_count_rows": status_count_rows,
        "statement_count": statements.count(),
        "amount_totals": amount_totals,
    }
    return render(request, "distribution/statement_center.html", context)


def _report_filters(request):
    today = timezone.localdate()
    default_start = today.replace(month=1, day=1)
    date_from = request.GET.get("date_from") or default_start.isoformat()
    date_to = request.GET.get("date_to") or today.isoformat()
    title_id = request.GET.get("title") or ""
    counterparty_id = request.GET.get("counterparty") or ""
    return {
        "date_from": date_from,
        "date_to": date_to,
        "title_id": title_id,
        "counterparty_id": counterparty_id,
    }


def _filtered_sales_reports(filters):
    reports = SalesReport.objects.select_related("title", "counterparty", "territory").filter(
        period_start__gte=filters["date_from"],
        period_start__lte=filters["date_to"],
    )
    if filters["title_id"]:
        reports = reports.filter(title_id=filters["title_id"])
    if filters["counterparty_id"]:
        reports = reports.filter(counterparty_id=filters["counterparty_id"])
    return reports


def _filtered_costs(filters):
    costs = Cost.objects.select_related("title", "supplier").filter(
        cost_date__gte=filters["date_from"],
        cost_date__lte=filters["date_to"],
    )
    if filters["title_id"]:
        costs = costs.filter(title_id=filters["title_id"])
    if filters["counterparty_id"]:
        costs = costs.filter(supplier_id=filters["counterparty_id"])
    return costs


def _filtered_waterfall_runs(filters):
    runs = WaterfallRun.objects.select_related("plan", "plan__title", "finalized_by").filter(
        period_start__lte=filters["date_to"],
        period_end__gte=filters["date_from"],
    )
    if filters["title_id"]:
        runs = runs.filter(plan__title_id=filters["title_id"])
    if filters["counterparty_id"]:
        runs = runs.filter(lines__beneficiary_id=filters["counterparty_id"]).distinct()
    return runs


@login_required
def reports(request):
    filters = _report_filters(request)
    sales_reports = _filtered_sales_reports(filters)
    costs = _filtered_costs(filters)
    net_expr = ExpressionWrapper(
        F("gross_revenue") - F("deductions") - F("vat_withholding"),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )

    revenue_by_title = (
        sales_reports.values("title__title_pl", "currency")
        .annotate(gross=Sum("gross_revenue"), net=Sum(net_expr), reports=Count("id"))
        .order_by("-net", "title__title_pl")
    )
    revenue_by_field = (
        sales_reports.values("exploitation_field", "currency")
        .annotate(gross=Sum("gross_revenue"), net=Sum(net_expr), reports=Count("id"))
        .order_by("-net", "exploitation_field")
    )
    revenue_by_counterparty = (
        sales_reports.values("counterparty__name", "currency")
        .annotate(gross=Sum("gross_revenue"), net=Sum(net_expr), reports=Count("id"))
        .order_by("-net", "counterparty__name")
    )
    costs_by_title = (
        costs.values("title__title_pl", "currency")
        .annotate(net=Sum("net_amount"), gross=Sum(F("net_amount") + F("vat_amount")), costs=Count("id"))
        .order_by("-net", "title__title_pl")
    )

    report_totals = []
    for currency in sorted(set(sales_reports.values_list("currency", flat=True)) | set(costs.values_list("currency", flat=True))):
        currency_sales = sales_reports.filter(currency=currency)
        currency_costs = costs.filter(currency=currency)
        gross_revenue = currency_sales.aggregate(total=Sum("gross_revenue"))["total"] or 0
        deductions = currency_sales.aggregate(total=Sum("deductions"))["total"] or 0
        vat_withholding = currency_sales.aggregate(total=Sum("vat_withholding"))["total"] or 0
        net_revenue = currency_sales.aggregate(total=Sum(net_expr))["total"] or 0
        cost_total = currency_costs.aggregate(total=Sum("net_amount"))["total"] or 0
        report_totals.append({"currency": currency, "gross_revenue": gross_revenue, "deductions": deductions, "vat_withholding": vat_withholding, "net_revenue": net_revenue, "cost_total": cost_total, "margin": net_revenue - cost_total})
    waterfall_runs = _filtered_waterfall_runs(filters).order_by("-period_end", "plan__title__title_pl")
    waterfall_totals = waterfall_runs.aggregate(
        gross_revenue=Sum("gross_revenue"),
        net_revenue=Sum("net_revenue"),
        allocated_amount=Sum("allocated_amount"),
        closing_available=Sum("closing_available"),
    )

    context = {
        "filters": filters,
        "titles": Title.objects.order_by("title_pl"),
        "counterparties": Counterparty.objects.order_by("name"),
        "report_totals": report_totals,
        "sales_count": sales_reports.count(),
        "cost_count": costs.count(),
        "revenue_by_title": revenue_by_title,
        "revenue_by_field": revenue_by_field,
        "revenue_by_counterparty": revenue_by_counterparty,
        "costs_by_title": costs_by_title,
        "waterfall_runs": waterfall_runs,
        "waterfall_totals": waterfall_totals,
        "overdue_agreements": SalesAgreement.objects.filter(invoice_paid=False, payment_due_date__lt=timezone.localdate()).select_related("title", "licensee")[:20],
        "open_issues": RightsIssue.objects.filter(resolved=False).select_related("rights_window", "rights_window__title")[:20],
        "catalog_export_form": TitleCatalogExportForm(),
    }
    return render(request, "distribution/reports.html", context)


@login_required
@require_POST
def title_catalog_export(request):
    form = TitleCatalogExportForm(request.POST)
    if not form.is_valid():
        errors = " ".join(error for field_errors in form.errors.values() for error in field_errors)
        messages.error(request, f"Nie można przygotować Eksportu 360. {errors}")
        return redirect("distribution:reports")

    if form.cleaned_data["export_all"]:
        title_ids = list(Title.objects.values_list("pk", flat=True))
    else:
        title_ids = list(form.cleaned_data["title_ids"].values_list("pk", flat=True))

    sheets = build_catalog_export(
        title_ids,
        date_from=form.cleaned_data["date_from"],
        date_to=form.cleaned_data["date_to"],
    )
    stamp = timezone.localdate().isoformat()
    if form.cleaned_data["export_format"] == "csv_zip":
        response = HttpResponse(export_catalog_csv_zip(sheets), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="filmerp_export_360_{stamp}.zip"'
        return response

    response = HttpResponse(
        export_catalog_xlsx(sheets),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="filmerp_export_360_{stamp}.xlsx"'
    return response


@login_required
def reports_export_csv(request):
    filters = _report_filters(request)
    export_format = request.GET.get("format", "csv")
    if export_format == "xlsx":
        return _reports_export_xlsx(filters)
    if request.GET.get("section") == "waterfall":
        return _waterfall_export_csv(filters)
    sales_reports = _filtered_sales_reports(filters).order_by("period_start", "title__title_pl")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="sales_report_export.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow([
        "title",
        "counterparty",
        "exploitation_field",
        "territory",
        "period_start",
        "period_end",
        "currency",
        "gross_revenue",
        "deductions",
        "vat_withholding",
        "net_revenue",
        "source_reference",
    ])
    for report in sales_reports:
        writer.writerow([
            report.title.title_pl,
            report.counterparty.name,
            report.get_exploitation_field_display(),
            report.territory.name if report.territory else "",
            report.period_start,
            report.period_end,
            report.currency,
            report.gross_revenue,
            report.deductions,
            report.vat_withholding,
            report.net_revenue,
            report.source_reference,
        ])
    return response


def _waterfall_export_csv(filters):
    waterfall_runs = _filtered_waterfall_runs(filters).order_by("period_start", "plan__title__title_pl")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="waterfall_period_settlements.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow([
        "title",
        "plan",
        "plan_version",
        "period_start",
        "period_end",
        "status",
        "currency",
        "gross_revenue",
        "net_revenue",
        "allocated_amount",
        "closing_available",
        "calculated_at",
        "finalized_at",
    ])
    for run in waterfall_runs:
        writer.writerow([
            run.plan.title.title_pl,
            run.plan.name,
            run.plan.version,
            run.period_start,
            run.period_end,
            run.get_status_display(),
            run.plan.currency,
            run.gross_revenue,
            run.net_revenue,
            run.allocated_amount,
            run.closing_available,
            run.calculated_at or "",
            run.finalized_at or "",
        ])
    return response


def _append_table(sheet, headers, rows):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF7")
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _statement_center_export_xlsx(statements, filters):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    statement_list = list(statements)
    _append_table(
        summary,
        ["Metric", "Value"],
        [
            ["Status", filters["status"] or "all"],
            ["Title", filters["title_id"] or "all"],
            ["Recipient", filters["recipient_id"] or "all"],
            ["Date from", filters["date_from"] or ""],
            ["Date to", filters["date_to"] or ""],
            ["Statement count", len(statement_list)],
            ["Amount due total", sum((statement.amount_due for statement in statement_list), 0)],
            ["Open amount total", sum((statement.amount_due for statement in statement_list if statement.status != StatementStatus.PAID), 0)],
        ],
    )
    details = workbook.create_sheet("Statements")
    _append_table(
        details,
        ["Title", "Recipient", "Period start", "Period end", "Currency", "Gross", "Net", "Recoupable costs", "Distributor fee", "Net receipts", "Amount due", "Status", "Sent at", "Paid at", "PDF"],
        [
            [
                statement.title.title_pl,
                statement.recipient.name,
                statement.period_start,
                statement.period_end,
                statement.currency,
                statement.gross_revenue,
                statement.net_revenue,
                statement.recoupable_costs,
                statement.distributor_fee_amount,
                statement.net_receipts,
                statement.amount_due,
                statement.get_status_display(),
                statement.sent_at or "",
                statement.paid_at or "",
                statement.statement_file.url if statement.statement_file else "",
            ]
            for statement in statement_list
        ],
    )
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="filmerp_statement_center.xlsx"'
    return response


def _reports_export_xlsx(filters):
    sales_reports = _filtered_sales_reports(filters).order_by("period_start", "title__title_pl")
    waterfall_runs = list(_filtered_waterfall_runs(filters).order_by("period_start", "plan__title__title_pl"))
    costs = _filtered_costs(filters).order_by("cost_date", "title__title_pl")
    overdue_agreements = SalesAgreement.objects.filter(invoice_paid=False, payment_due_date__lt=timezone.localdate()).select_related("title", "licensee")
    open_issues = RightsIssue.objects.filter(resolved=False).select_related("rights_window", "rights_window__title")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary_rows = [
        ["Date from", filters["date_from"]],
        ["Date to", filters["date_to"]],
        ["Sales report count", sales_reports.count()],
        ["Cost count", costs.count()],
        ["Waterfall period settlement count", len(waterfall_runs)],
    ]
    _append_table(summary, ["Metric", "Value"], summary_rows)

    sales_sheet = workbook.create_sheet("Sales reports")
    _append_table(
        sales_sheet,
        ["Title", "Counterparty", "Field", "Territory", "Period start", "Period end", "Currency", "Gross", "Deductions", "VAT withholding", "Net", "Source"],
        [
            [
                report.title.title_pl,
                report.counterparty.name,
                report.get_exploitation_field_display(),
                report.territory.name if report.territory else "",
                report.period_start,
                report.period_end,
                report.currency,
                report.gross_revenue,
                report.deductions,
                report.vat_withholding,
                report.net_revenue,
                report.source_reference,
            ]
            for report in sales_reports
        ],
    )

    waterfall_sheet = workbook.create_sheet("Waterfall")
    _append_table(
        waterfall_sheet,
        ["Title", "Plan", "Version", "Period start", "Period end", "Status", "Currency", "Gross revenue", "Net revenue", "Allocated", "Remaining", "Calculated at", "Finalized at"],
        [
            [
                run.plan.title.title_pl,
                run.plan.name,
                run.plan.version,
                run.period_start,
                run.period_end,
                run.get_status_display(),
                run.plan.currency,
                run.gross_revenue,
                run.net_revenue,
                run.allocated_amount,
                run.closing_available,
                run.calculated_at.isoformat() if run.calculated_at else "",
                run.finalized_at.isoformat() if run.finalized_at else "",
            ]
            for run in waterfall_runs
        ],
    )

    costs_sheet = workbook.create_sheet("Costs")
    _append_table(
        costs_sheet,
        ["Title", "Supplier", "Category", "Scope mode", "Scope", "Allocation percentages", "Date", "Currency", "Net", "VAT", "Gross", "Recoupable", "Recovered", "Outstanding", "Paid"],
        [
            [
                cost.title.title_pl,
                cost.supplier.name if cost.supplier else "",
                cost.get_category_display(),
                cost.get_scope_mode_display(),
                cost.scope_label,
                ", ".join(f"{field}: {percentage}%" for field, percentage in cost.allocation_percentages.items()),
                cost.cost_date,
                cost.currency,
                cost.net_amount,
                cost.vat_amount,
                cost.gross_amount,
                "yes" if cost.recoupable else "no",
                cost.recouped_amount,
                cost.outstanding_recoupment,
                "yes" if cost.paid else "no",
            ]
            for cost in costs
        ],
    )

    alerts_sheet = workbook.create_sheet("Alerts")
    _append_table(
        alerts_sheet,
        ["Type", "Title", "Counterparty", "Date", "Message"],
        [[
            "Overdue payment",
            agreement.title.title_pl,
            agreement.licensee.name,
            agreement.payment_due_date,
            f"{agreement.fixed_fee} {agreement.currency}",
        ] for agreement in overdue_agreements] + [[
            "Rights issue",
            issue.rights_window.title.title_pl,
            issue.rights_window.counterparty.name if issue.rights_window.counterparty else "",
            issue.created_at.date(),
            issue.message,
        ] for issue in open_issues],
    )

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="filmerp_report_export.xlsx"'
    return response
